"""Полный GORJI Excel импорт (задача 4.2.1).

**Стратегия Variant B (прагматичная, упрощённая):**
- DASH cols D..AT копируем в absolute period_number=k без shift'а.
- Launch lag через `psc.launch_year/launch_month` (D-13 механизм) — pipeline
  сам обнулит nd/offtake до launch периода.
- Logistics берём константой M1 (без per-period inflation в pipeline).
- Material/Package — базовое M1, инфляция через `inflate_series` (как в Discovery V1).
- Project-level CAPEX/OPEX из DATA rows 33/26 → ProjectFinancialPlan
  записи на первом периоде каждого года.
- Seasonality NULL — не привязываем, чтобы избежать double application
  (DASH ND/offtake уже могут содержать seasonal pattern).

**Известные расхождения** (измеряются acceptance test'ом в конце):

1. **DASH относительная ось не shift'нута.** Excel хранит DASH cols
   относительно launch month канала, а в листах NET REVENUE / VOLUME
   сдвигает в absolute periods. Мы копируем напрямую → значения в
   первые активные periods после launch будут "плато ramp" вместо
   "начала ramp". Завышение output в эти periods.
2. **Logistics не инфлируется в pipeline.** Excel инфлирует logistics
   с тем же профилем что shelf price. Наш pipeline хранит как float
   константу. Расхождение: +7% logistics стоимости с Y2 (2025+) → лёгкое
   занижение Contribution.

Запуск:
    docker compose -f infra/docker-compose.dev.yml exec backend \\
        python -m scripts.import_gorji_full

Требования:
    - PASSPORT_MODEL_GORJI_2025-09-05.xlsx скопирован в /tmp/gorji.xlsx
      контейнера backend (`docker cp ... backend:/tmp/gorji.xlsx`)
    - openpyxl установлен (`pip install openpyxl` в контейнере)
"""
from __future__ import annotations

import asyncio
from datetime import date
from decimal import Decimal
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import delete, select

from app.db import async_session_maker
from app.models import (
    BOMItem,
    Channel,
    Period,
    PeriodType,
    PeriodValue,
    Project,
    ProjectFinancialPlan,
    ProjectSKU,
    ProjectSKUChannel,
    RefInflation,
    RefSeasonality,
    SKU,
    Scenario,
    ScenarioResult,
    ScenarioType,
    SourceType,
)
from app.models.base import PeriodScope
from app.schemas.project import ProjectCreate
from app.schemas.project_sku_channel import ProjectSKUChannelCreate
from app.services.calculation_service import calculate_all_scenarios
from app.services.project_service import create_project
from app.services.project_sku_channel_service import create_psk_channel

# ============================================================
# Constants — Excel layout (см. зонды _probe_dash*.py от 2026-04-09)
# ============================================================

XLSX_PATH = "/tmp/gorji.xlsx"
PROJECT_NAME = "GORJI+ полный импорт (4.2.1)"

# 8 SKU блоков, шаг 46 строк
SKU_BASE_ROWS: list[int] = [6, 52, 98, 144, 190, 236, 282, 328]

# 6 каналов через col offset. label col = col_base, value col = col_base + 1
# (для launch_year/month), period cols = col_base + 2 .. col_base + 44 (43 значения)
CHANNEL_LAYOUT: list[tuple[str, int]] = [
    ("HM", 2),
    ("SM", 50),
    ("MM", 98),
    ("TT", 146),
    ("E-COM_OZ", 194),
    ("E-COM_OZ_Fresh", 242),
]

# Row offsets от base SKU блока
OFFSET_LAUNCH_YEAR = 2
OFFSET_LAUNCH_MONTH = 3
OFFSET_BRAND = 8
OFFSET_NAME = 9
OFFSET_PACKAGE = 11
OFFSET_VOLUME = 14
OFFSET_ND = 19              # per-period (col_base+2..+44)
OFFSET_OFFTAKE = 20         # per-period
OFFSET_CHANNEL_MARGIN = 21  # col_base+2 (M1, константа)
OFFSET_PROMO_DISCOUNT = 22  # col_base+2
OFFSET_PROMO_SHARE = 23     # col_base+2
OFFSET_SHELF_PRICE = 24     # per-period
OFFSET_MATERIAL = 30        # per-period (M1 = база, остальное инфляция)
OFFSET_PACKAGE_COST = 31    # per-period
OFFSET_PROD_RATE = 32       # col_base+2 (одинаков для всех каналов SKU)
OFFSET_LOGISTIC = 34        # per-period (берём M1 как константу)
OFFSET_CAM_RATE = 35        # col_base+2
OFFSET_MARKETING_RATE = 36  # col_base+2

PERIOD_COL_START_OFFSET = 2  # col_base + 2 = первая period col
PERIOD_COUNT = 43            # M1..M36 (36) + Y4..Y10 (7)

# Project-level из rows 1-5 DASH (хедер выше первого SKU блока)
PROJECT_HEADER_ROW_INFLATION = 2   # col C
PROJECT_HEADER_ROW_DR = 3          # col C (wacc)
PROJECT_HEADER_ROW_VAT = 22        # col C (внутри SKU 1 блока)

# DATA лист — KPI эталоны и project-level CAPEX/OPEX
DATA_ROW_OPEX = 26      # B..K
DATA_ROW_CAPEX = 33     # B..K
DATA_ROW_NPV = 48       # B/C/D = Y1Y3 / Y1Y5 / Y1Y10
DATA_ROW_ROI = 49
DATA_ROW_IRR = 50
DATA_ROW_PAYBACK_SIMPLE = 51
DATA_ROW_PAYBACK_DISC = 52


# ============================================================
# Excel extraction
# ============================================================


def _dec(v: Any) -> Decimal:
    """Конвертация Excel float → Decimal с защитой от None."""
    if v is None:
        return Decimal("0")
    return Decimal(str(v))


def _float(v: Any, default: float = 0.0) -> float:
    if v is None:
        return default
    return float(v)


def extract_project_header(wb) -> dict[str, Any]:
    """Project-level: inflation, wacc, vat, start_date, horizon."""
    dash = wb["DASH"]
    return {
        "start_date": date(2024, 1, 1),  # B1=C1 = 2024-01-01
        "horizon_years": 10,
        "wacc": _dec(dash.cell(PROJECT_HEADER_ROW_DR, 3).value),
        "tax_rate": Decimal("0.20"),  # дефолт по ADR-CE-04
        "wc_rate": Decimal("0.12"),   # дефолт по ADR-CE-02
        "vat_rate": _dec(dash.cell(PROJECT_HEADER_ROW_VAT, 3).value),
        "inflation_profile_name": dash.cell(
            PROJECT_HEADER_ROW_INFLATION, 3
        ).value,
    }


def extract_sku_block(wb, sku_idx: int) -> dict[str, Any]:
    """Извлекает один SKU блок (метаданные SKU + 6 каналов)."""
    dash = wb["DASH"]
    base = SKU_BASE_ROWS[sku_idx]

    # SKU метаданные читаем из value col канала HM (col_base+1=3)
    hm_value_col = CHANNEL_LAYOUT[0][1] + 1  # = 3 (C)
    hm_period_col_m1 = CHANNEL_LAYOUT[0][1] + PERIOD_COL_START_OFFSET  # = 4 (D)

    sku_meta = {
        "brand": dash.cell(base + OFFSET_BRAND, hm_value_col).value,
        "name": dash.cell(base + OFFSET_NAME, hm_value_col).value,
        "package_type": dash.cell(base + OFFSET_PACKAGE, hm_value_col).value,
        "volume_l": dash.cell(base + OFFSET_VOLUME, hm_value_col).value,
    }

    # Per-SKU rates (одинаковы для всех каналов, читаем из M1 col HM)
    # D-19: production_cost_rate теперь per-period в DASH (Excel switches
    # off rate during copacking window). Static значение в ProjectSKU
    # остаётся как fallback для UI и проектов без per-period override.
    # Берём максимум из ряда per-period rates как "стандартное" значение.
    prod_rate_raw = dash.cell(base + OFFSET_PROD_RATE, hm_period_col_m1).value
    if not prod_rate_raw:  # SKU 7-8 имеют 0 в M1 col, ищем максимум по периодам
        max_rate = 0.0
        for c in range(hm_period_col_m1, hm_period_col_m1 + 43):
            v = dash.cell(base + OFFSET_PROD_RATE, c).value or 0
            if v > max_rate:
                max_rate = v
        prod_rate_raw = max_rate

    sku_rates = {
        "production_cost_rate": _dec(prod_rate_raw),
        "ca_m_rate": _dec(
            dash.cell(base + OFFSET_CAM_RATE, hm_period_col_m1).value
        ),
        "marketing_rate": _dec(
            dash.cell(base + OFFSET_MARKETING_RATE, hm_period_col_m1).value
        ),
    }

    # BOM (per-SKU, читаем из M1 col HM — берём базовые M1 значения,
    # pipeline применит инфляцию через inflate_series)
    bom = {
        "material_cost_m1": _dec(
            dash.cell(base + OFFSET_MATERIAL, hm_period_col_m1).value
        ),
        "package_cost_m1": _dec(
            dash.cell(base + OFFSET_PACKAGE_COST, hm_period_col_m1).value
        ),
    }

    # 6 каналов
    channels = []
    for code, col_base in CHANNEL_LAYOUT:
        value_col = col_base + 1
        m1_col = col_base + PERIOD_COL_START_OFFSET

        # DASH cols 4..39 = relative monthly M1..M36 (36 значений)
        # DASH cols 40..46 = relative yearly Y4..Y10 (7 значений)
        # Извлекаем raw — shift применим в import_to_db (нужен launch_period)
        nd_monthly = [
            _float(dash.cell(base + OFFSET_ND, m1_col + i).value)
            for i in range(36)
        ]
        nd_yearly = [
            _float(dash.cell(base + OFFSET_ND, m1_col + 36 + i).value)
            for i in range(7)
        ]
        offtake_monthly = [
            _float(dash.cell(base + OFFSET_OFFTAKE, m1_col + i).value)
            for i in range(36)
        ]
        offtake_yearly = [
            _float(dash.cell(base + OFFSET_OFFTAKE, m1_col + 36 + i).value)
            for i in range(7)
        ]
        shelf_monthly = [
            _float(dash.cell(base + OFFSET_SHELF_PRICE, m1_col + i).value)
            for i in range(36)
        ]
        shelf_yearly = [
            _float(dash.cell(base + OFFSET_SHELF_PRICE, m1_col + 36 + i).value)
            for i in range(7)
        ]
        # Material/Package/Logistic per-period (D-16, D-18: per-period values
        # с custom Excel inflation, не из inflate_series)
        material_monthly = [
            _float(dash.cell(base + OFFSET_MATERIAL, m1_col + i).value)
            for i in range(36)
        ]
        material_yearly = [
            _float(dash.cell(base + OFFSET_MATERIAL, m1_col + 36 + i).value)
            for i in range(7)
        ]
        package_monthly = [
            _float(dash.cell(base + OFFSET_PACKAGE_COST, m1_col + i).value)
            for i in range(36)
        ]
        package_yearly = [
            _float(dash.cell(base + OFFSET_PACKAGE_COST, m1_col + 36 + i).value)
            for i in range(7)
        ]
        logistic_monthly = [
            _float(dash.cell(base + OFFSET_LOGISTIC, m1_col + i).value)
            for i in range(36)
        ]
        logistic_yearly = [
            _float(dash.cell(base + OFFSET_LOGISTIC, m1_col + 36 + i).value)
            for i in range(7)
        ]
        # D-20: channel_margin / promo_discount / promo_share per-period
        # (Excel меняет promo_share с 1.0 в M1..M27 до 0.8 в Y4..Y10)
        cm_monthly = [
            _float(dash.cell(base + OFFSET_CHANNEL_MARGIN, m1_col + i).value)
            for i in range(36)
        ]
        cm_yearly = [
            _float(dash.cell(base + OFFSET_CHANNEL_MARGIN, m1_col + 36 + i).value)
            for i in range(7)
        ]
        pd_monthly = [
            _float(dash.cell(base + OFFSET_PROMO_DISCOUNT, m1_col + i).value)
            for i in range(36)
        ]
        pd_yearly = [
            _float(dash.cell(base + OFFSET_PROMO_DISCOUNT, m1_col + 36 + i).value)
            for i in range(7)
        ]
        ps_monthly = [
            _float(dash.cell(base + OFFSET_PROMO_SHARE, m1_col + i).value)
            for i in range(36)
        ]
        ps_yearly = [
            _float(dash.cell(base + OFFSET_PROMO_SHARE, m1_col + 36 + i).value)
            for i in range(7)
        ]
        # D-19: production_cost_rate per-period (Excel переключает rate
        # по периодам — copacking window для own production downtime)
        prod_rate_monthly = [
            _float(dash.cell(base + OFFSET_PROD_RATE, m1_col + i).value)
            for i in range(36)
        ]
        prod_rate_yearly = [
            _float(dash.cell(base + OFFSET_PROD_RATE, m1_col + 36 + i).value)
            for i in range(7)
        ]

        ch_data = {
            "code": code,
            "launch_year_excel": _float(
                dash.cell(base + OFFSET_LAUNCH_YEAR, value_col).value
            ),
            "launch_month_excel": _float(
                dash.cell(base + OFFSET_LAUNCH_MONTH, value_col).value
            ),
            "channel_margin": _dec(
                dash.cell(base + OFFSET_CHANNEL_MARGIN, m1_col).value
            ),
            "promo_discount": _dec(
                dash.cell(base + OFFSET_PROMO_DISCOUNT, m1_col).value
            ),
            "promo_share": _dec(
                dash.cell(base + OFFSET_PROMO_SHARE, m1_col).value
            ),
            "shelf_price_m1": _dec(
                dash.cell(base + OFFSET_SHELF_PRICE, m1_col).value
            ),
            "logistics_m1": _dec(
                dash.cell(base + OFFSET_LOGISTIC, m1_col).value
            ),
            "nd_target": max(nd_monthly + nd_yearly) if nd_monthly else 0.0,
            "offtake_target": max(offtake_monthly + offtake_yearly) if offtake_monthly else 0.0,
            # Raw monthly + yearly relative — shift'ятся в import_to_db
            "nd_monthly": nd_monthly,
            "nd_yearly": nd_yearly,
            "offtake_monthly": offtake_monthly,
            "offtake_yearly": offtake_yearly,
            "shelf_monthly": shelf_monthly,
            "shelf_yearly": shelf_yearly,
            "material_monthly": material_monthly,
            "material_yearly": material_yearly,
            "package_monthly": package_monthly,
            "package_yearly": package_yearly,
            "logistic_monthly": logistic_monthly,
            "logistic_yearly": logistic_yearly,
            "cm_monthly": cm_monthly,
            "cm_yearly": cm_yearly,
            "pd_monthly": pd_monthly,
            "pd_yearly": pd_yearly,
            "ps_monthly": ps_monthly,
            "ps_yearly": ps_yearly,
            "prod_rate_monthly": prod_rate_monthly,
            "prod_rate_yearly": prod_rate_yearly,
        }
        channels.append(ch_data)

    return {
        "sku_idx": sku_idx + 1,
        "meta": sku_meta,
        "rates": sku_rates,
        "bom": bom,
        "channels": channels,
    }


def extract_project_capex_opex(wb) -> dict[int, tuple[float, float]]:
    """Извлекает yearly CAPEX/OPEX из DATA rows 33/26 + copacking r22.

    Маппинг календарь → наш period_number:
        Y0=2024 → M1 (period 1)
        Y1=2025 → M13
        Y2=2026 → M25
        Y3=2027 → period 37 (Y4)
        Y4=2028 → period 38 (Y5)
        ...
        Y9=2033 → period 43 (Y10)

    D-21: Excel DATA r22 (Копакинг) показывает 6.96M в Y1=2025 и нули в
    остальные годы. Это launch-год copacking затрат, которые в Excel
    вычитаются из GP (subtraction в COGS_total). У нас pipeline не имеет
    per-period copacking source — добавляем как extra OPEX к Y1=2025
    period 13 (effect на Contribution идентичен effect на GP в Excel).

    Возвращает {period_number: (capex, opex)}.
    """
    data = wb["DATA"]
    DATA_ROW_COPACKING = 22  # Excel DATA r22 "Копакинг, ₽"
    # cols B..K = 2..11 = 10 годов (Y0..Y9)
    # period_number маппинг: B(2)→1, C(3)→13, D(4)→25, E(5)→37, F→38, ..., K→43
    excel_col_to_period = {
        2: 1,    # Y0=2024 → M1
        3: 13,   # Y1=2025 → M13
        4: 25,   # Y2=2026 → M25
        5: 37,   # Y3=2027 → Y4
        6: 38,   # Y4=2028 → Y5
        7: 39,   # Y5=2029 → Y6
        8: 40,   # Y6=2030 → Y7
        9: 41,   # Y7=2031 → Y8
        10: 42,  # Y8=2032 → Y9
        11: 43,  # Y9=2033 → Y10
    }
    plan: dict[int, tuple[float, float]] = {}
    for excel_col, period_number in excel_col_to_period.items():
        capex = _float(data.cell(DATA_ROW_CAPEX, excel_col).value)
        opex = _float(data.cell(DATA_ROW_OPEX, excel_col).value)
        # D-21: добавляем copacking в opex (effect на Contribution = effect
        # на GP, общая идея — снижение FCF этого года на сумму copacking)
        copacking = _float(data.cell(DATA_ROW_COPACKING, excel_col).value)
        opex += copacking
        plan[period_number] = (capex, opex)
    return plan


def extract_kpi_reference(wb) -> dict[str, dict[str, Any]]:
    """KPI эталоны из DATA rows 48/49/50/51/52 для трёх scope."""
    data = wb["DATA"]
    return {
        "y1y3": {
            "npv": _float(data.cell(DATA_ROW_NPV, 2).value),
            "roi": _float(data.cell(DATA_ROW_ROI, 2).value),
            "irr": _float(data.cell(DATA_ROW_IRR, 2).value),
            "payback_simple": data.cell(DATA_ROW_PAYBACK_SIMPLE, 2).value,
            "payback_discounted": data.cell(DATA_ROW_PAYBACK_DISC, 2).value,
        },
        "y1y5": {
            "npv": _float(data.cell(DATA_ROW_NPV, 3).value),
            "roi": _float(data.cell(DATA_ROW_ROI, 3).value),
            "irr": _float(data.cell(DATA_ROW_IRR, 3).value),
            "payback_simple": data.cell(DATA_ROW_PAYBACK_SIMPLE, 3).value,
            "payback_discounted": data.cell(DATA_ROW_PAYBACK_DISC, 3).value,
        },
        "y1y10": {
            "npv": _float(data.cell(DATA_ROW_NPV, 4).value),
            "roi": _float(data.cell(DATA_ROW_ROI, 4).value),
            "irr": _float(data.cell(DATA_ROW_IRR, 4).value),
            "payback_simple": data.cell(DATA_ROW_PAYBACK_SIMPLE, 4).value,
            "payback_discounted": data.cell(DATA_ROW_PAYBACK_DISC, 4).value,
        },
    }


# ============================================================
# Excel year/month → наш launch_year/launch_month
# ============================================================


def excel_to_project_launch(
    excel_year: int, excel_month: int, project_start: date
) -> tuple[int, int]:
    """Конвертирует Excel calendar year/month в наши launch_year (1..10) +
    launch_month (1..12) относительно project.start_date.

    Excel year=2024, month=11 + start=2024-01-01 → launch_year=1, launch_month=11
    Excel year=2025, month=2 + start=2024 → launch_year=2, launch_month=2
    """
    launch_year = excel_year - project_start.year + 1
    return launch_year, int(excel_month)


def launch_to_period_number(launch_year: int, launch_month: int) -> int:
    """Маппинг launch_year/month → absolute period_number (1..43).

    Y1 Jan = period 1, Y1 Dec = 12, Y2 Jan = 13, ..., Y3 Dec = 36
    Y4 = 37, Y5 = 38, ..., Y10 = 43
    """
    if launch_year <= 3:
        return (launch_year - 1) * 12 + launch_month
    return 36 + (launch_year - 3)


def shift_dash_to_absolute(
    monthly_vals: list[float],   # 36 values: DASH cols 4..39 = relative M1..M36 канала
    yearly_vals: list[float],    # 7 values: DASH cols 40..46 = relative Y4..Y10 канала
    launch_period: int,          # 1..43 absolute period где DASH M1 канала
    launch_year: int,            # 1..10 наш проектный launch_year
) -> list[float]:
    """Shift DASH relative-to-launch values в absolute project periods (D-15).

    DASH хранит значения относительно launch month канала:
    - cols 4..39 = monthly relative M1..M36 канала (3 года ramp + плато)
    - cols 40..46 = yearly relative Y4..Y10 канала

    Наш проект имеет 36 monthly absolute (M1..M36) + 7 yearly absolute (Y4..Y10).

    Algorithm:
    1. DASH monthly col i (i=0..35) → absolute_period = launch_period + i
       - Если absolute_period <= 36 (зона monthly absolute): copy directly
       - Если absolute_period > 36 (попадает в yearly absolute): aggregate
         в нужный absolute year (average по cols, попадающим в этот год)
    2. DASH yearly col i (i=0..6, представляет relative Y(i+4) канала) →
       absolute_year = launch_year + (i + 4) - 1
       - Если 4 <= absolute_year <= 10: добавить в abs period (37 + abs_year - 4)
       - Иначе: за horizon, игнорируется

    Все наши per-period values — **rates** (nd, offtake_per_outlet, shelf_price,
    material_per_unit, package_per_unit, logistic_per_kg). При сворачивании
    multiple cols в один yearly absolute period — **average** (не sum).

    Returns: list 43 значения в absolute axis. Periods до launch и за horizon = 0.
    """
    n_periods = 43
    abs_vals = [0.0] * n_periods
    abs_counts = [0] * n_periods

    # 1) DASH monthly cols → absolute periods
    for i, val in enumerate(monthly_vals):
        abs_period = launch_period + i  # 1-indexed
        if abs_period > n_periods + 30:  # совсем за horizon
            break
        if abs_period <= 36:
            # Direct copy в monthly absolute
            abs_vals[abs_period - 1] = val
            abs_counts[abs_period - 1] = 1
        else:
            # Falls into absolute yearly. abs_period 37..49+ → abs_year 4..N
            abs_year = (abs_period - 37) // 12 + 4  # M37..M48 = Y4, M49..M60 = Y5, ...
            if 4 <= abs_year <= 10:
                abs_period_yearly = 36 + (abs_year - 3)  # Y4 → 37, Y10 → 43
                abs_vals[abs_period_yearly - 1] += val
                abs_counts[abs_period_yearly - 1] += 1
            # else: за horizon

    # 2) DASH yearly cols → absolute yearly periods
    for i, val in enumerate(yearly_vals):
        relative_year = i + 4  # 4..10 (DASH col 40..46)
        abs_year = launch_year + relative_year - 1
        if 4 <= abs_year <= 10:
            abs_period = 36 + (abs_year - 3)  # 37..43
            abs_vals[abs_period - 1] += val
            abs_counts[abs_period - 1] += 1

    # 3) Average для periods где count > 1 (multiple sources contributed)
    for j in range(n_periods):
        if abs_counts[j] > 1:
            abs_vals[j] /= abs_counts[j]

    return abs_vals


# ============================================================
# Cleanup существующего проекта (hard delete cascade)
# ============================================================


async def cleanup_existing_project(session, name: str) -> None:
    """Удаляет проект с этим именем (включая soft-deleted), CASCADE."""
    # Hard delete — нужно убрать deleted_at filter
    rows = (
        await session.execute(
            select(Project).where(Project.name == name)
        )
    ).scalars().all()
    for p in rows:
        print(f"  Cleaning up existing project id={p.id} (deleted_at={p.deleted_at})")
        await session.delete(p)
    await session.flush()


# ============================================================
# Import to DB
# ============================================================


async def import_to_db(
    session,
    header: dict,
    sku_blocks: list[dict],
    capex_opex: dict[int, tuple[float, float]],
) -> int:
    """Создаёт project + 8 SKU + 8 ProjectSKU + 48 PSC + 6192 PeriodValue +
    10 ProjectFinancialPlan записей.

    Возвращает project.id.
    """
    # 1. Inflation profile lookup
    inflation = await session.scalar(
        select(RefInflation).where(
            RefInflation.profile_name == header["inflation_profile_name"]
        )
    )
    if inflation is None:
        raise RuntimeError(
            f"Inflation profile {header['inflation_profile_name']!r} "
            "не найден в seed_reference_data — нужно засеять или проверить имя"
        )

    # 1b. Seasonality WTR profile (GORJI Category = WTR из DASH B10)
    wtr_seasonality = await session.scalar(
        select(RefSeasonality).where(RefSeasonality.profile_name == "WTR")
    )
    wtr_seasonality_id = wtr_seasonality.id if wtr_seasonality else None

    # 2. Project
    project = await create_project(
        session,
        ProjectCreate(
            name=PROJECT_NAME,
            start_date=header["start_date"],
            horizon_years=header["horizon_years"],
            wacc=header["wacc"],
            tax_rate=header["tax_rate"],
            wc_rate=header["wc_rate"],
            vat_rate=header["vat_rate"],
            currency="RUB",
            inflation_profile_id=inflation.id,
        ),
    )
    await session.flush()
    print(f"  Created Project id={project.id}: {project.name}")
    print(f"    wacc={project.wacc} vat={project.vat_rate} "
          f"tax={project.tax_rate} wc={project.wc_rate}")

    # 3. Periods catalog (отсортированы)
    periods = list(
        (
            await session.scalars(
                select(Period).order_by(Period.period_number)
            )
        ).all()
    )
    assert len(periods) == PERIOD_COUNT, f"Expected {PERIOD_COUNT} periods, got {len(periods)}"

    # Scenarios для PeriodValue (создаются автоматически в create_project)
    scenarios = list(
        (
            await session.scalars(
                select(Scenario).where(Scenario.project_id == project.id)
            )
        ).all()
    )
    assert len(scenarios) == 3, f"Expected 3 scenarios, got {len(scenarios)}"

    # Channels lookup (один раз)
    channels_by_code: dict[str, Channel] = {}
    for code, _ in CHANNEL_LAYOUT:
        ch = await session.scalar(select(Channel).where(Channel.code == code))
        if ch is None:
            raise RuntimeError(f"Channel {code!r} not in seed")
        channels_by_code[code] = ch

    # 4. 8 SKU + 8 ProjectSKU + 48 PSC + 6192 PeriodValue
    total_psc = 0
    total_pv = 0
    for block in sku_blocks:
        meta = block["meta"]
        rates = block["rates"]
        bom = block["bom"]

        # SKU справочник
        sku = SKU(
            brand=meta["brand"],
            name=meta["name"],
            format=meta["package_type"],
            volume_l=_dec(meta["volume_l"]),
            package_type=meta["package_type"],
        )
        session.add(sku)
        await session.flush()

        # ProjectSKU с rates
        psk = ProjectSKU(
            project_id=project.id,
            sku_id=sku.id,
            production_cost_rate=rates["production_cost_rate"],
            ca_m_rate=rates["ca_m_rate"],
            marketing_rate=rates["marketing_rate"],
        )
        session.add(psk)
        await session.flush()

        # BOM (2 строки на SKU: material + package, базовые M1)
        session.add(
            BOMItem(
                project_sku_id=psk.id,
                ingredient_name="Material (GORJI)",
                quantity_per_unit=Decimal("1"),
                loss_pct=Decimal("0"),
                price_per_unit=bom["material_cost_m1"],
            )
        )
        session.add(
            BOMItem(
                project_sku_id=psk.id,
                ingredient_name="Package (GORJI)",
                quantity_per_unit=Decimal("1"),
                loss_pct=Decimal("0"),
                price_per_unit=bom["package_cost_m1"],
            )
        )
        await session.flush()

        # 6 ProjectSKUChannel
        for ch_data in block["channels"]:
            channel = channels_by_code[ch_data["code"]]

            # Excel year/month → launch_year/month
            launch_year, launch_month = excel_to_project_launch(
                int(ch_data["launch_year_excel"]),
                int(ch_data["launch_month_excel"]),
                header["start_date"],
            )
            launch_period = launch_to_period_number(launch_year, launch_month)

            # D-15 ОТМЕНЁН: DASH cells оказались **absolute** (не relative).
            # Discovery V2: Volume Y4 точно совпадал с Excel БЕЗ shift'а
            # (3,694,359 = 3,694,359). NR_M1 = 0 в NET REVENUE объясняется
            # тем что Excel применяет launch_lag в листах NET REVENUE/VOLUME,
            # обнуляя periods до launch month. DASH cells содержат "planned
            # values" без launch lag, и pipeline launch lag (D-13 механизм)
            # обнуляет до launch_period — что соответствует Excel поведению.
            # Прямой copy DASH cols без shift = absolute axis.
            nd_abs = ch_data["nd_monthly"] + ch_data["nd_yearly"]
            offtake_abs = ch_data["offtake_monthly"] + ch_data["offtake_yearly"]
            shelf_abs = ch_data["shelf_monthly"] + ch_data["shelf_yearly"]
            material_abs = ch_data["material_monthly"] + ch_data["material_yearly"]
            package_abs = ch_data["package_monthly"] + ch_data["package_yearly"]
            logistic_abs = ch_data["logistic_monthly"] + ch_data["logistic_yearly"]
            cm_abs = ch_data["cm_monthly"] + ch_data["cm_yearly"]
            pd_abs = ch_data["pd_monthly"] + ch_data["pd_yearly"]
            ps_abs = ch_data["ps_monthly"] + ch_data["ps_yearly"]
            prod_rate_abs = ch_data["prod_rate_monthly"] + ch_data["prod_rate_yearly"]

            # Создаём PSC БЕЗ auto_fill_predict (мы пишем свои PREDICT)
            psc = await create_psk_channel(
                session,
                psk.id,
                ProjectSKUChannelCreate(
                    channel_id=channel.id,
                    launch_year=launch_year,
                    launch_month=launch_month,
                    nd_target=_dec(ch_data["nd_target"]),
                    nd_ramp_months=12,  # для UI; pipeline использует predict values
                    offtake_target=_dec(ch_data["offtake_target"]),
                    channel_margin=ch_data["channel_margin"],
                    promo_discount=ch_data["promo_discount"],
                    promo_share=ch_data["promo_share"],
                    shelf_price_reg=ch_data["shelf_price_m1"],
                    logistics_cost_per_kg=ch_data["logistics_m1"],
                    seasonality_profile_id=wtr_seasonality_id,  # WTR (Category из DASH)
                ),
                auto_fill_predict=False,  # сами пишем PREDICT ниже
            )
            total_psc += 1

            # Записываем PREDICT PeriodValue × 3 сценария × 43 периода
            # Это 129 записей на канал × 48 каналов = 6192 записей.
            #
            # D-16, D-17, D-18: значения material/package/logistic/shelf
            # копируем из DASH cells (после D-15 shift) — это содержит
            # Excel custom inflation, не наш inflate_series.
            for scenario in scenarios:
                for i, period in enumerate(periods):
                    pv = PeriodValue(
                        psk_channel_id=psc.id,
                        scenario_id=scenario.id,
                        period_id=period.id,
                        source_type=SourceType.PREDICT,
                        version_id=1,
                        is_overridden=False,
                        values={
                            "nd": nd_abs[i],
                            "offtake": offtake_abs[i],
                            "shelf_price": shelf_abs[i],
                            # D-16: material+package per-period (combined для bom_unit_cost)
                            "bom_unit_cost": material_abs[i] + package_abs[i],
                            # D-18: logistics per-period
                            "logistic_per_kg": logistic_abs[i],
                            # D-20: channel_margin / promo per-period
                            "channel_margin": cm_abs[i],
                            "promo_discount": pd_abs[i],
                            "promo_share": ps_abs[i],
                            # D-19: production_cost_rate per-period
                            "production_cost_rate": prod_rate_abs[i],
                        },
                    )
                    session.add(pv)
                    total_pv += 1

        # Flush после каждого SKU чтобы не накапливать огромный pending state
        await session.flush()
        print(
            f"  SKU {block['sku_idx']} '{meta['name'][:35]}': "
            f"PSK id={psk.id}, 6 каналов, {6 * 3 * 43} PeriodValue"
        )

    print(f"\n  Total PSC: {total_psc}, Total PeriodValue: {total_pv}")

    # 5. ProjectFinancialPlan — 10 yearly записей
    print("\n  Creating ProjectFinancialPlan...")
    period_by_number = {p.period_number: p for p in periods}
    for period_number, (capex, opex) in capex_opex.items():
        period = period_by_number[period_number]
        plan = ProjectFinancialPlan(
            project_id=project.id,
            period_id=period.id,
            capex=Decimal(str(capex)),
            opex=Decimal(str(opex)),
        )
        session.add(plan)
        print(
            f"    period_number={period_number} ({period.type.value}, "
            f"model_year={period.model_year}): capex={capex:>14,.0f} opex={opex:>14,.0f}"
        )
    await session.flush()

    return project.id


# ============================================================
# Acceptance: сравнить ScenarioResult с эталоном
# ============================================================


def _fmt_money(v) -> str:
    if v is None:
        return "—"
    return f"{float(v):>15,.0f}₽"


def _fmt_pct(v) -> str:
    if v is None:
        return "—"
    return f"{float(v) * 100:>7.2f}%"


def _diff_pct(actual: float | None, expected: float) -> str:
    if actual is None:
        return "N/A"
    if expected == 0:
        return f"{actual - expected:+,.2f} (abs)"
    delta = (float(actual) - expected) / abs(expected) * 100
    return f"{delta:+.2f}%"


async def diagnostic_annual_aggregates(session, project_id: int) -> None:
    """Печатает annual NR / Contribution / FCF Base сценария для diagnosис.

    Excel эталоны (DATA rows 22-43, cols B..K = Y0..Y9):
        Y0(Y1): GP=145,572      CM=108,151      FCF=-6,546,718
        Y1(Y2): GP=12,764,110   CM=-103,999     FCF=-10,183,030
        Y2(Y3): GP=42,380,597   CM=23,206,801   FCF=4,971,324
        Y3(Y4): GP=65,179,798   CM=39,154,289   FCF=19,414,536
        Y4(Y5): GP=75,435,427   CM=45,700,563   FCF=27,400,693
        Y5(Y6): GP=86,981,103   CM=53,059,271   FCF=32,597,354
        Y6(Y7): GP=100,031,512  CM=61,390,128   FCF=38,503,715
        Y7(Y8): GP=114,759,719  CM=70,805,412   FCF=45,211,635
        Y8(Y9): GP=131,357,043  CM=81,429,261   FCF=52,814,368
        Y9(Y10): GP=146,504,177 CM=91,098,561   FCF=60,586,701
    """
    from app.engine.aggregator import aggregate_lines
    from app.engine.pipeline import run_line_pipeline
    from app.services.calculation_service import (
        _load_period_catalog,
        _load_project_financial_plan,
        build_line_inputs,
    )
    from app.engine.pipeline import run_project_pipeline

    print("\n" + "=" * 78)
    print("DIAGNOSTIC: annual aggregates Base сценария vs Excel")
    print("=" * 78)

    base_scenario = await session.scalar(
        select(Scenario).where(
            Scenario.project_id == project_id,
            Scenario.type == ScenarioType.BASE,
        )
    )

    inputs = await build_line_inputs(session, project_id, base_scenario.id)
    sorted_periods, _ = await _load_period_catalog(session)
    capex, opex = await _load_project_financial_plan(
        session, project_id, sorted_periods
    )

    agg = run_project_pipeline(inputs, project_capex=capex, project_opex=opex)

    # Per-line aggregator перед s10 — там есть annual GP/COGS/Logistics
    # перед свёрткой в monthly aggregate. Запустим вручную s01..s09 +
    # aggregator чтобы получить monthly GP/COGS/Logistics для diagnostics.
    line_contexts = [run_line_pipeline(inp) for inp in inputs]
    pre_aggregate = aggregate_lines(
        line_contexts, project_capex=capex, project_opex=opex
    )
    # Annualize per-period в годовые корзины
    annual_gp_buf: dict[int, float] = {}
    annual_cogs_buf: dict[int, float] = {}
    annual_logistics_buf: dict[int, float] = {}
    annual_camk_buf: dict[int, float] = {}  # CA&M + Marketing
    annual_mat_buf: dict[int, float] = {}
    annual_prod_buf: dict[int, float] = {}
    annual_volu_buf: dict[int, float] = {}
    for t in range(pre_aggregate.input.period_count):
        year = pre_aggregate.input.period_model_year[t]
        annual_gp_buf[year] = annual_gp_buf.get(year, 0) + pre_aggregate.gross_profit[t]
        annual_cogs_buf[year] = annual_cogs_buf.get(year, 0) + pre_aggregate.cogs_total[t]
        annual_logistics_buf[year] = annual_logistics_buf.get(year, 0) + pre_aggregate.logistics_cost[t]
        annual_mat_buf[year] = annual_mat_buf.get(year, 0) + pre_aggregate.cogs_material[t]
        annual_prod_buf[year] = annual_prod_buf.get(year, 0) + pre_aggregate.cogs_production[t]
        annual_volu_buf[year] = annual_volu_buf.get(year, 0) + pre_aggregate.volume_units[t]

    print(f"\n  COGS breakdown по годам (наш vs Excel):")
    print(f"  {'Year':6}{'Volume':>15}{'Mat (наш)':>15}{'Mat+Pkg Excel':>17}"
          f"{'Prod (наш)':>15}{'Prod Excel':>15}")
    excel_mat_pkg_ref = [   # Excel rows 19+20 (Material + Package)
        34990 + 56849, 7365192 + 11581838, 21160027 + 33307162,
        30115581 + 47397452, 34849693 + 54845612, 40212261 + 63282272,
        46244500 + 72822422, 53181175 + 83593269, 61158351 + 95735458,
        68097533 + 106809663,
    ]
    excel_prod_ref = [20018, 246393, 8190046, 12065168, 13961923,
                      16104389, 18526599, 21260762, 24342493, 27154033]
    excel_units_ref = [8171, 1173706, 2954371, 3694359, 3990656,
                       4298361, 4631038, 4989564, 5375698, 5731006]  # DATA r15
    for y in sorted(annual_gp_buf.keys())[:10]:
        idx = y - 1
        v_ours = annual_volu_buf[y]
        v_ex = excel_units_ref[idx] if idx < 10 else 0
        m_ours = annual_mat_buf[y]
        m_ex = excel_mat_pkg_ref[idx] if idx < 10 else 0
        p_ours = annual_prod_buf[y]
        p_ex = excel_prod_ref[idx] if idx < 10 else 0
        v_diff = (v_ours - v_ex) / v_ex * 100 if v_ex else 0
        print(f"  Y{y:<5d}{v_ours:>15,.0f}{m_ours:>15,.0f}{m_ex:>17,.0f}"
              f"{p_ours:>15,.0f}{p_ex:>15,.0f}  vol_Δ={v_diff:+6.1f}%")
    print()
    print(f"  GP/Logistics breakdown:")
    print(f"  {'Year':6}{'Excel GP':>15}{'Наш GP':>15}{'Δ%':>8}"
          f"{'Excel Logist':>15}{'Наш Logist':>15}{'Δ%':>8}")
    excel_gp_ref = [145572, 12764110, 42380597, 65179798, 75435427,
                    86981103, 100031512, 114759719, 131357043, 146504177]
    excel_logistics_ref = [37420, 9648108, 15792796, 22475459, 26007311,
                           30007902, 34531757, 39639199, 45396918, 50648209]
    for y in sorted(annual_gp_buf.keys())[:10]:
        idx = y - 1
        gp_ours = annual_gp_buf[y]
        log_ours = annual_logistics_buf[y]
        gp_ex = excel_gp_ref[idx]
        log_ex = excel_logistics_ref[idx]
        gp_diff = (gp_ours - gp_ex) / gp_ex * 100 if gp_ex else 0
        log_diff = (log_ours - log_ex) / log_ex * 100 if log_ex else 0
        print(f"  Y{y:<5d}{gp_ex:>15,.0f}{gp_ours:>15,.0f}{gp_diff:>7.1f}%"
              f"{log_ex:>15,.0f}{log_ours:>15,.0f}{log_diff:>7.1f}%")
    print()

    # Excel эталоны для сравнения
    excel_gp = excel_gp_ref
    excel_cm = [108151, -103999, 23206801, 39154289, 45700563,
                53059271, 61390128, 70805412, 81429261, 91098561]
    excel_fcf = [-6546718, -10183030, 4971324, 19414536, 27400693,
                 32597354, 38503715, 45211635, 52814368, 60586701]
    excel_capex = [6602348, 5440000, 5659500, 5942475, 6239599,
                   6551579, 6879158, 7223116, 7584271, 7963485]
    excel_opex = [0, 3220000, 3381000, 3550050, 3727553,
                  3913930, 4109627, 4315108, 4530863, 4757407]

    print(f"\n  {'Year':6}{'Excel NR/CM/FCF':>50s}{'Наш NR/CM/FCF':>50s}")
    print("  " + "-" * 110)
    print(f"  {'':6}{'NR':>15}{'CM':>15}{'FCF':>20s}{'NR':>15}{'CM':>15}{'FCF':>20s}")
    n_years = len(agg.annual_net_revenue)
    for y in range(min(10, n_years)):
        ours_nr = agg.annual_net_revenue[y]
        ours_cm = agg.annual_contribution[y]
        ours_fcf = agg.annual_free_cash_flow[y]
        ex_cm = excel_cm[y] if y < 10 else None
        ex_fcf = excel_fcf[y] if y < 10 else None
        print(
            f"  Y{y + 1:<5d}"
            f"{'':>15s}"  # Excel NR не доступен в DATA
            f"{ex_cm:>15,.0f}"
            f"{ex_fcf:>20,.0f}"
            f"{ours_nr:>15,.0f}"
            f"{ours_cm:>15,.0f}"
            f"{ours_fcf:>20,.0f}"
        )

    # Total
    total_ours_fcf = sum(agg.annual_free_cash_flow)
    total_excel_fcf = sum(excel_fcf)
    print(f"\n  Total FCF: наш = {total_ours_fcf:>15,.0f}  vs  "
          f"Excel = {total_excel_fcf:>15,.0f}  ratio = {total_ours_fcf/total_excel_fcf:.3f}")

    # Discounted
    print(f"\n  Annual DCF (наш):")
    for y, dcf in enumerate(agg.annual_discounted_cash_flow):
        print(f"    Y{y+1}: {dcf:>15,.0f}")
    print(f"  Sum DCF Y1Y10: {sum(agg.annual_discounted_cash_flow):>15,.0f}")
    print(f"  Excel NPV Y1Y10: 79,983,059")


async def acceptance_compare(
    session, project_id: int, kpi_ref: dict
) -> dict[str, dict]:
    """Сравнивает ScenarioResult Base сценария с эталоном из DATA."""
    print("\n" + "=" * 78)
    print("ACCEPTANCE: сравнение Base сценария с GORJI Excel эталоном")
    print("=" * 78)

    base_scenario = await session.scalar(
        select(Scenario).where(
            Scenario.project_id == project_id,
            Scenario.type == ScenarioType.BASE,
        )
    )
    assert base_scenario is not None

    results = (
        await session.scalars(
            select(ScenarioResult).where(
                ScenarioResult.scenario_id == base_scenario.id
            )
        )
    ).all()
    by_scope = {r.period_scope.value: r for r in results}

    diffs: dict[str, dict] = {}
    for scope in ("y1y3", "y1y5", "y1y10"):
        ref = kpi_ref[scope]
        actual = by_scope.get(scope)
        if actual is None:
            print(f"\n  Scope {scope}: НЕТ результата")
            continue

        actual_npv = float(actual.npv) if actual.npv is not None else None
        actual_irr = float(actual.irr) if actual.irr is not None else None
        actual_roi = float(actual.roi) if actual.roi is not None else None

        print(f"\n  Scope {scope.upper()}:")
        print(f"    {'KPI':10s}{'Excel эталон':>20s}{'Наш расчёт':>20s}{'Δ':>15s}")
        print(f"    {'-' * 65}")
        print(f"    {'NPV':10s}{_fmt_money(ref['npv']):>20s}"
              f"{_fmt_money(actual_npv):>20s}"
              f"{_diff_pct(actual_npv, ref['npv']):>15s}")
        print(f"    {'IRR':10s}{_fmt_pct(ref['irr']):>20s}"
              f"{_fmt_pct(actual_irr):>20s}"
              f"{_diff_pct(actual_irr, ref['irr']):>15s}")
        print(f"    {'ROI':10s}{_fmt_pct(ref['roi']):>20s}"
              f"{_fmt_pct(actual_roi):>20s}"
              f"{_diff_pct(actual_roi, ref['roi']):>15s}")
        print(f"    {'Payback s':10s}{str(ref['payback_simple']):>20s}"
              f"{str(actual.payback_simple):>20s}")
        print(f"    {'Payback d':10s}{str(ref['payback_discounted']):>20s}"
              f"{str(actual.payback_discounted):>20s}")

        diffs[scope] = {
            "npv_actual": actual_npv,
            "npv_expected": ref["npv"],
            "irr_actual": actual_irr,
            "irr_expected": ref["irr"],
            "roi_actual": actual_roi,
            "roi_expected": ref["roi"],
        }

    return diffs


# ============================================================
# Main
# ============================================================


async def main() -> None:
    print("=" * 78)
    print("GORJI+ полный импорт (4.2.1, Variant B упрощённый)")
    print("=" * 78)
    print()

    print("Loading Excel...")
    wb = load_workbook(XLSX_PATH, data_only=True)

    print("Extracting project header...")
    header = extract_project_header(wb)
    print(f"  start_date={header['start_date']} horizon={header['horizon_years']}")
    print(f"  wacc={header['wacc']} vat={header['vat_rate']} tax={header['tax_rate']}")
    print(f"  inflation={header['inflation_profile_name']!r}")
    print()

    print("Extracting 8 SKU blocks...")
    sku_blocks = [extract_sku_block(wb, i) for i in range(8)]
    for b in sku_blocks:
        print(
            f"  SKU {b['sku_idx']}: {b['meta']['brand']} / "
            f"{str(b['meta']['name'])[:50]} ({b['meta']['volume_l']}L)"
        )
        for ch in b["channels"]:
            ly_excel = int(ch["launch_year_excel"])
            lm_excel = int(ch["launch_month_excel"])
            ly_proj, lm_proj = excel_to_project_launch(
                ly_excel, lm_excel, header["start_date"]
            )
            print(
                f"    {ch['code']:15s} launch_excel={ly_excel}/{lm_excel:02d} "
                f"→ launch_project=Y{ly_proj}/M{lm_proj:02d}  "
                f"cm={ch['channel_margin']} shelf_M1={ch['shelf_price_m1']}"
            )
    print()

    print("Extracting project CAPEX/OPEX from DATA...")
    capex_opex = extract_project_capex_opex(wb)
    print(f"  10 yearly entries")
    print()

    print("Extracting KPI reference from DATA...")
    kpi_ref = extract_kpi_reference(wb)
    print(f"  Y1-Y3:  NPV={kpi_ref['y1y3']['npv']:>15,.0f}  "
          f"IRR={kpi_ref['y1y3']['irr']*100:.2f}%  "
          f"ROI={kpi_ref['y1y3']['roi']*100:.2f}%")
    print(f"  Y1-Y5:  NPV={kpi_ref['y1y5']['npv']:>15,.0f}  "
          f"IRR={kpi_ref['y1y5']['irr']*100:.2f}%  "
          f"ROI={kpi_ref['y1y5']['roi']*100:.2f}%")
    print(f"  Y1-Y10: NPV={kpi_ref['y1y10']['npv']:>15,.0f}  "
          f"IRR={kpi_ref['y1y10']['irr']*100:.2f}%  "
          f"ROI={kpi_ref['y1y10']['roi']*100:.2f}%")
    print()

    wb.close()

    # === Импорт в БД ===
    print("=" * 78)
    print("Importing to database...")
    print("=" * 78)
    async with async_session_maker() as session:
        await cleanup_existing_project(session, PROJECT_NAME)

        project_id = await import_to_db(
            session, header, sku_blocks, capex_opex
        )
        await session.commit()
        print(f"\n  ✅ Import complete. Project id={project_id}")

        # === Recalculate ===
        print("\n" + "=" * 78)
        print("Running calculate_all_scenarios...")
        print("=" * 78)
        results_by_scenario = await calculate_all_scenarios(session, project_id)
        await session.commit()
        print(f"  ✅ Calculated {sum(len(r) for r in results_by_scenario.values())} "
              f"ScenarioResult rows")

        # === Diagnostic ===
        await diagnostic_annual_aggregates(session, project_id)

        # === Acceptance ===
        diffs = await acceptance_compare(session, project_id, kpi_ref)

        # === Verdict ===
        print("\n" + "=" * 78)
        print("VERDICT")
        print("=" * 78)
        max_npv_drift = 0.0
        for scope, d in diffs.items():
            if d.get("npv_actual") is not None and d["npv_expected"] != 0:
                drift = abs(
                    (d["npv_actual"] - d["npv_expected"]) / d["npv_expected"]
                )
                max_npv_drift = max(max_npv_drift, drift)

        print(f"  Max NPV drift: {max_npv_drift * 100:.2f}%")
        if max_npv_drift < 0.01:
            print("  ✅ ACCEPTANCE PASSED (drift < 1%)")
        elif max_npv_drift < 0.05:
            print("  ⚠️  ACCEPTANCE: drift 1-5% — приемлемо для Variant B")
        elif max_npv_drift < 0.15:
            print("  ⚠️  ACCEPTANCE: drift 5-15% — нужны исправления")
        else:
            print("  ❌ ACCEPTANCE: drift > 15% — Variant B не подходит, "
                  "переходить на A")
        print()


if __name__ == "__main__":
    asyncio.run(main())
