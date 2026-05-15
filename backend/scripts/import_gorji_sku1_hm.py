"""Discovery: импорт SKU_1/HM из GORJI Excel через pipeline (Phase 4.2.1).

**Цель:** проверить что `build_line_inputs + run_project_pipeline` работают
на реальных GORJI данных, а не синтетических. Закрытие долга задачи 2.4 —
end-to-end acceptance с реальным wiring от БД до KPI.

One-shot скрипт. Не в CI. Требует openpyxl установленный вручную:
    docker cp PASSPORT_MODEL_GORJI_2025-09-05.xlsx \\
        dbpassport-dev-backend-1:/tmp/gorji.xlsx
    docker compose exec backend pip install openpyxl
    docker compose exec backend python -m scripts.import_gorji_sku1_hm

Извлекает ВВОДНЫЕ данные из первого SKU блока DASH (SKU_1 × HM):
- Project: wacc, vat, inflation profile
- SKU: brand, name, format, volume_l
- ProjectSKU: production_cost_rate (ca_m/marketing — на PSC после Q6, 2026-05-15)
- BOM: material + package cost (2 BOMItem)
- PSC HM: nd_target, offtake_target, channel_margin, promo_*, shelf_price_reg, logistics
- Per-period nd/offtake/shelf_price для 43 периодов (пишем как finetuned
  поверх auto-fill predict, приоритет finetuned > predict)

Затем:
1. Создаёт проект с именем "GORJI+ SKU1/HM discovery"
2. Запускает calculate_all_scenarios напрямую
3. Печатает KPI по 3 сценариям × 3 скоупам
4. Сравнивает per-unit GP с DASH row 44 (эталон 14.43 ₽/unit M1-M3, 13.74 M4-M6)

Verdict выводится в конце — "МАТЧ" или "РАСХОЖДЕНИЕ".
"""
from __future__ import annotations

import asyncio
from datetime import date
from decimal import Decimal
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select

from app.db import async_session_maker
from app.engine.context import PipelineInput
from app.engine.pipeline import run_line_pipeline
from app.models import (
    BOMItem,
    Channel,
    Period,
    PeriodValue,
    Project,
    ProjectSKU,
    RefInflation,
    SKU,
    Scenario,
    ScenarioType,
    SourceType,
)
from app.schemas.project import ProjectCreate
from app.schemas.project_sku_channel import ProjectSKUChannelCreate
from app.services.calculation_service import (
    build_line_inputs,
    calculate_all_scenarios,
)
from app.services.project_service import create_project
from app.services.project_sku_channel_service import create_psk_channel

XLSX_PATH = "/tmp/gorji.xlsx"
PROJECT_NAME = "GORJI+ SKU1/HM discovery"

# Эталонные per-unit значения из DASH row 44 (GP) и 46 (CM)
# M1-M3 до апрельской инфляции, M4-M6 после
EXPECTED_GP_PER_UNIT_M1 = 14.429289012939108
EXPECTED_GP_PER_UNIT_M4 = 13.744954327931834


def extract_first_sku_hm(xlsx_path: str) -> dict[str, Any]:
    """Извлекает данные первого SKU блока DASH (SKU_1 × HM).

    Структура DASH (первый блок, rows 6-48):
    - row 14: brand
    - row 15: SKU name
    - row 17: package type
    - row 20: volume l
    - row 22 col C: VAT %
    - row 2 col C: inflation profile name
    - row 12 col C: channel code
    - rows 25/26/30: ND/offtake/shelf_price per period (cols D..AT = 4..46)
    - row 27/28/29 col D: channel_margin/promo_discount/promo_share
    - row 36 col D: material cost per unit (M1, до инфляции)
    - row 37 col D: package cost per unit
    - row 38 col D: production_cost_rate
    - row 40 col D: logistics_cost_per_kg (M1)
    - row 41 col D: ca_m_rate
    - row 42 col D: marketing_rate
    """
    wb = load_workbook(xlsx_path, data_only=True)
    dash = wb["DASH"]

    # Project-level
    vat_rate = dash.cell(22, 3).value
    inflation_name = dash.cell(2, 3).value

    # SKU
    brand = dash.cell(14, 3).value
    sku_name = dash.cell(15, 3).value
    package_type = dash.cell(17, 3).value
    volume_l = dash.cell(20, 3).value

    # Channel
    channel_code = dash.cell(12, 3).value

    # PSC static params (col D = M1, параметры константные)
    channel_margin = dash.cell(27, 4).value
    promo_discount = dash.cell(28, 4).value
    promo_share = dash.cell(29, 4).value

    # ProjectSKU rates (col D, константные)
    production_cost_rate = dash.cell(38, 4).value
    ca_m_rate = dash.cell(41, 4).value
    marketing_rate = dash.cell(42, 4).value

    # BOM unit cost (col D = M1 base, до инфляции)
    material_cost = dash.cell(36, 4).value
    package_cost = dash.cell(37, 4).value

    # Shelf price base (col D = M1)
    shelf_price_base = dash.cell(30, 4).value
    # Logistics cost (col D = M1 base, до инфляции)
    logistics_cost = dash.cell(40, 4).value

    # Targets — максимум по всем периодам
    # Cols D..AT = 4..46 (43 периода)
    nd_values = [dash.cell(25, c).value or 0 for c in range(4, 47)]
    offtake_values = [dash.cell(26, c).value or 0 for c in range(4, 47)]
    shelf_values = [dash.cell(30, c).value or 0 for c in range(4, 47)]

    nd_target = max(nd_values)
    offtake_target = max(offtake_values)

    # Per-period (для finetuned override)
    periods_data = [
        {
            "period_number": i + 1,
            "nd": nd_values[i],
            "offtake": offtake_values[i],
            "shelf_price": shelf_values[i],
        }
        for i in range(43)
    ]

    wb.close()

    return {
        "vat_rate": vat_rate,
        "inflation_profile_name": inflation_name,
        "brand": brand,
        "sku_name": sku_name,
        "package_type": package_type,
        "volume_l": volume_l,
        "channel_code": channel_code,
        "channel_margin": channel_margin,
        "promo_discount": promo_discount,
        "promo_share": promo_share,
        "production_cost_rate": production_cost_rate,
        "ca_m_rate": ca_m_rate,
        "marketing_rate": marketing_rate,
        "material_cost": material_cost,
        "package_cost": package_cost,
        "shelf_price_base": shelf_price_base,
        "logistics_cost": logistics_cost,
        "nd_target": nd_target,
        "offtake_target": offtake_target,
        "periods": periods_data,
    }


def _dec(v: Any) -> Decimal:
    return Decimal(str(v)) if v is not None else Decimal("0")


async def cleanup_existing_project(session) -> None:
    existing = await session.scalar(
        select(Project).where(Project.name == PROJECT_NAME)
    )
    if existing is not None:
        print(f"  Deleting existing project id={existing.id}")
        await session.delete(existing)
        await session.flush()


async def main() -> None:
    print("=== GORJI+ SKU1/HM Discovery Import ===\n")
    print("Extracting data from Excel...")
    data = extract_first_sku_hm(XLSX_PATH)

    print(f"  SKU: {data['brand']} / {data['sku_name']}")
    print(f"  Channel: {data['channel_code']}")
    print(f"  Volume: {data['volume_l']} L")
    print(f"  ND target: {data['nd_target']}")
    print(f"  Offtake target: {data['offtake_target']}")
    print(f"  Shelf price base: {data['shelf_price_base']} ₽")
    print(f"  Material+Package: {data['material_cost']:.4f} + {data['package_cost']:.4f}")
    print(f"  Production rate: {data['production_cost_rate']:.6f}")
    print(f"  CA&M rate: {data['ca_m_rate']:.6f}")
    print(f"  Marketing rate: {data['marketing_rate']:.6f}")
    print(f"  Inflation: {data['inflation_profile_name']}")
    print()

    async with async_session_maker() as session:
        await cleanup_existing_project(session)

        # 1. Find inflation profile
        inflation = await session.scalar(
            select(RefInflation).where(
                RefInflation.profile_name == data["inflation_profile_name"]
            )
        )
        if inflation is None:
            print(f"  ⚠ Inflation profile '{data['inflation_profile_name']}' not in seed")
        inflation_id = inflation.id if inflation is not None else None

        # 2. Create project
        project = await create_project(
            session,
            ProjectCreate(
                name=PROJECT_NAME,
                start_date=date(2025, 1, 1),
                horizon_years=10,
                wacc=Decimal("0.19"),
                tax_rate=Decimal("0.20"),
                wc_rate=Decimal("0.12"),
                vat_rate=_dec(data["vat_rate"]),
                currency="RUB",
                inflation_profile_id=inflation_id,
            ),
        )
        print(f"  Created Project id={project.id}")

        # 3. Create SKU
        sku = SKU(
            brand=data["brand"],
            name=data["sku_name"],
            format=data["package_type"],
            volume_l=_dec(data["volume_l"]),
            package_type=data["package_type"],
        )
        session.add(sku)
        await session.flush()
        print(f"  Created SKU id={sku.id}")

        # 4. Create ProjectSKU
        # Q6 (2026-05-15): ca_m_rate и marketing_rate переехали на ProjectSKUChannel.
        psk = ProjectSKU(
            project_id=project.id,
            sku_id=sku.id,
            production_cost_rate=_dec(data["production_cost_rate"]),
        )
        session.add(psk)
        await session.flush()
        print(f"  Created ProjectSKU id={psk.id}")

        # 5. Create BOM items — Material и Package отдельными строками
        session.add(
            BOMItem(
                project_sku_id=psk.id,
                ingredient_name="Material (из GORJI)",
                quantity_per_unit=Decimal("1"),
                loss_pct=Decimal("0"),
                price_per_unit=_dec(data["material_cost"]),
            )
        )
        session.add(
            BOMItem(
                project_sku_id=psk.id,
                ingredient_name="Package (из GORJI)",
                quantity_per_unit=Decimal("1"),
                loss_pct=Decimal("0"),
                price_per_unit=_dec(data["package_cost"]),
            )
        )
        await session.flush()
        print(f"  Created 2 BOMItem (material + package)")

        # 6. Create PSC HM
        hm = await session.scalar(
            select(Channel).where(Channel.code == data["channel_code"])
        )
        if hm is None:
            raise RuntimeError(
                f"Channel {data['channel_code']!r} not in seed — проверь seed_reference_data"
            )

        psc = await create_psk_channel(
            session,
            psk.id,
            ProjectSKUChannelCreate(
                channel_id=hm.id,
                nd_target=_dec(data["nd_target"]),
                offtake_target=_dec(data["offtake_target"]),
                channel_margin=_dec(data["channel_margin"]),
                promo_discount=_dec(data["promo_discount"]),
                promo_share=_dec(data["promo_share"]),
                shelf_price_reg=_dec(data["shelf_price_base"]),
                logistics_cost_per_kg=_dec(data["logistics_cost"]),
                # Q6 (2026-05-15): CA&M/Marketing per-channel.
                ca_m_rate=_dec(data["ca_m_rate"]),
                marketing_rate=_dec(data["marketing_rate"]),
            ),
            auto_fill_predict=True,  # создаст 129 PeriodValue predict
        )
        print(f"  Created PSC id={psc.id} (auto-fill predict)")

        # 7. Override с finetuned per-period values из Excel
        periods = list(
            (
                await session.scalars(
                    select(Period).order_by(Period.period_number)
                )
            ).all()
        )
        scenarios = list(
            (
                await session.scalars(
                    select(Scenario).where(Scenario.project_id == project.id)
                )
            ).all()
        )

        # Finetuned для всех 3 сценариев (delta пока = 0)
        for scenario in scenarios:
            for period, pdata in zip(periods, data["periods"]):
                session.add(
                    PeriodValue(
                        psk_channel_id=psc.id,
                        scenario_id=scenario.id,
                        period_id=period.id,
                        source_type=SourceType.FINETUNED,
                        version_id=1,
                        is_overridden=True,
                        values={
                            "nd": float(pdata["nd"]),
                            "offtake": float(pdata["offtake"]),
                            "shelf_price": float(pdata["shelf_price"]),
                        },
                    )
                )
        await session.flush()
        print(f"  Created {len(periods) * len(scenarios)} finetuned PeriodValue overrides")

        await session.commit()

        # ========================================
        # 8. Per-line verification через build_line_inputs
        # ========================================
        print("\n=== Per-line verification (DASH row 44/46) ===")
        base_scenario = next(s for s in scenarios if s.type == ScenarioType.BASE)
        inputs = await build_line_inputs(session, project.id, base_scenario.id)
        assert len(inputs) == 1, f"Expected 1 line, got {len(inputs)}"
        inp = inputs[0]

        # Прогоняем s01..s09 напрямую
        ctx = run_line_pipeline(inp)

        # Сверка M1 (period index 0) per-unit значений
        for t, label in [(0, "M1"), (1, "M2"), (2, "M3"), (3, "M4"), (4, "M5"), (5, "M6")]:
            if ctx.volume_units[t] > 0:
                gp_per_unit = ctx.gross_profit[t] / ctx.volume_units[t]
                cm_per_unit = ctx.contribution[t] / ctx.volume_units[t]
                print(
                    f"  {label}: GP/unit={gp_per_unit:7.4f}₽  CM/unit={cm_per_unit:7.4f}₽  "
                    f"vol={ctx.volume_units[t]:.2f} nr={ctx.net_revenue[t]:.2f}"
                )

        # Проверка against эталон
        gp_m1 = ctx.gross_profit[0] / ctx.volume_units[0]
        gp_m4 = ctx.gross_profit[3] / ctx.volume_units[3]
        m1_ok = abs(gp_m1 - EXPECTED_GP_PER_UNIT_M1) / EXPECTED_GP_PER_UNIT_M1 < 1e-4
        m4_ok = abs(gp_m4 - EXPECTED_GP_PER_UNIT_M4) / EXPECTED_GP_PER_UNIT_M4 < 1e-4

        print()
        print(f"  Эталон M1 GP/unit: {EXPECTED_GP_PER_UNIT_M1:.6f}₽")
        print(f"  Наш M1 GP/unit:    {gp_m1:.6f}₽  {'✓ МАТЧ' if m1_ok else '✗ РАСХОЖДЕНИЕ'}")
        print(f"  Эталон M4 GP/unit: {EXPECTED_GP_PER_UNIT_M4:.6f}₽")
        print(f"  Наш M4 GP/unit:    {gp_m4:.6f}₽  {'✓ МАТЧ' if m4_ok else '✗ РАСХОЖДЕНИЕ'}")
        print()

        # ========================================
        # 9. Full calculate_all_scenarios → ScenarioResult
        # ========================================
        print("=== Full calculate_all_scenarios ===")
        results = await calculate_all_scenarios(session, project.id)
        await session.commit()

        for sc_id, res_list in results.items():
            sc = await session.get(Scenario, sc_id)
            print(f"\nScenario {sc.type.value}:")
            for r in res_list:
                npv = float(r.npv) if r.npv is not None else None
                irr = float(r.irr) if r.irr is not None else None
                roi = float(r.roi) if r.roi is not None else None
                print(
                    f"  {r.period_scope.value:6s} "
                    f"NPV={npv:>18,.0f}₽ " if npv is not None else f"  {r.period_scope.value:6s} NPV=—            ",
                    end="",
                )
                print(
                    f" IRR={irr:>7.2%} ROI={roi:>7.2%} Go={r.go_no_go}"
                    if irr is not None and roi is not None
                    else f" Go={r.go_no_go}"
                )

        print()
        if m1_ok and m4_ok:
            print("🎉 DISCOVERY SUCCESS — per-line pipeline на реальных GORJI данных")
            print("   соответствует DASH эталону. Можем переходить к Варианту 2")
            print("   (полный импорт всех SKU × каналов).")
        else:
            print("❌ РАСХОЖДЕНИЕ per-unit значений — нужно расследовать:")
            print("   - Проверить маппинг колонок Excel (DASH row 25/26/30)")
            print("   - Проверить что все параметры psc/psk извлечены верно")
            print("   - Проверить что finetuned действительно используется в pipeline")


if __name__ == "__main__":
    asyncio.run(main())
