"""Генератор XLSX-экспорта проекта (задача 5.1, F-08).

Структура файла:
- Лист "Вводные": параметры проекта + таблица SKU + таблица каналов
- Лист "PnL по периодам": все per-period финансовые показатели Base сценария
- Лист "KPI": NPV/IRR/ROI/Payback × 3 сценария × 3 scope
- Лист "P&L Pivot": per-line breakdown (1 строка = SKU × Channel × Period)
  для пользовательских сводных таблиц в Excel (C #15)

Реализация через openpyxl (Apache 2.0). Pure Python, без pandas.
Возвращает `bytes` (in-memory XLSX), endpoint оборачивает в StreamingResponse.

Контракт:
- Service не ходит в БД сам — принимает уже загруженные ORM-объекты
  через async helpers (`load_project_data`).
- Pipeline аналог: `load_*` функции делают select, `build_workbook`
  работает с in-memory dataclass'ами.
"""
from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.engine.context import PipelineContext, PipelineInput
from app.engine.pipeline import run_line_pipeline, run_project_pipeline
from app.models import (
    BOMItem,
    Period,
    PeriodType,
    Project,
    ProjectFinancialPlan,
    ProjectSKU,
    ProjectSKUChannel,
    RefInflation,
    Scenario,
    ScenarioResult,
    ScenarioType,
)
from app.models.base import PeriodScope
from app.services.calculation_service import (
    _load_period_catalog,
    _load_project_financial_plan,
    build_line_inputs,
)


# ============================================================
# Style helpers
# ============================================================

HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUB_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
SUB_HEADER_FONT = Font(bold=True, size=10)
LABEL_FONT = Font(bold=True, size=10)


def _set_header(ws: Worksheet, row: int, headers: list[str]) -> None:
    """Заполняет строку заголовков с цветной заливкой и жирным шрифтом."""
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _set_section(ws: Worksheet, row: int, col: int, text: str) -> None:
    """Заголовок секции (мерж не нужен — просто bold)."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = SUB_HEADER_FILL
    cell.font = SUB_HEADER_FONT


def _autosize_columns(ws: Worksheet, max_width: int = 40) -> None:
    """Грубая авто-ширина по содержимому колонок."""
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for cell in ws[col_letter]:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            if length > max_len:
                max_len = length
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


# C #15: column headers for the P&L Pivot sheet (26 columns)
PNL_PIVOT_HEADERS: list[str] = [
    "Бренд",               # sku.brand
    "SKU",                 # sku.name
    "Формат",              # sku.format
    "Объём",               # sku.volume_l (numeric)
    "Единица",             # sku.unit_of_measure ("л" / "кг")
    "Код канала",          # channel.code
    "Канал",               # channel.name
    "Группа канала",       # channel.channel_group (C #16)
    "Источник",            # channel.source_type (C #16, nullable)
    "Период",              # "M1" / "Y4" / etc
    "Тип периода",         # "monthly" / "annual"
    "Год",                 # 1..10
    "Месяц",               # 1..12 / None
    "Квартал",             # 1..4 / None
    "Объём, ед",           # volume_units
    "Объём, л/кг",         # volume_liters
    "Чистая выручка, ₽",  # net_revenue
    "COGS сырьё, ₽",      # cogs_material
    "COGS произв., ₽",    # cogs_production
    "COGS итого, ₽",      # cogs_total
    "Gross Profit, ₽",    # gross_profit
    "Логистика, ₽",       # logistics_cost
    "Contribution, ₽",    # contribution
    "CA&M, ₽",            # ca_m_cost
    "Маркетинг, ₽",       # marketing_cost
    "EBITDA, ₽",          # ebitda
]


# ============================================================
# Data loading
# ============================================================


async def _load_project_full(
    session: AsyncSession, project_id: int
) -> Project | None:
    """Загружает Project + relationships (в одном запросе через selectinload)."""
    project = await session.get(Project, project_id)
    if project is None or project.deleted_at is not None:
        return None
    return project


async def _load_skus_with_bom(
    session: AsyncSession, project_id: int
) -> list[tuple[ProjectSKU, list[BOMItem]]]:
    """Загружает все ProjectSKU проекта с BOM items."""
    psk_rows = (
        await session.scalars(
            select(ProjectSKU)
            .where(ProjectSKU.project_id == project_id)
            .options(selectinload(ProjectSKU.sku))
        )
    ).all()

    result: list[tuple[ProjectSKU, list[BOMItem]]] = []
    for psk in psk_rows:
        bom = (
            await session.scalars(
                select(BOMItem).where(BOMItem.project_sku_id == psk.id)
            )
        ).all()
        result.append((psk, list(bom)))
    return result


async def _load_psk_channels(
    session: AsyncSession, project_id: int
) -> list[ProjectSKUChannel]:
    """Все ProjectSKUChannel проекта с подгруженными channel."""
    rows = (
        await session.scalars(
            select(ProjectSKUChannel)
            .join(ProjectSKU, ProjectSKU.id == ProjectSKUChannel.project_sku_id)
            .where(ProjectSKU.project_id == project_id)
            .options(
                selectinload(ProjectSKUChannel.channel),
            )
        )
    ).all()
    return list(rows)


async def _load_scenario_results(
    session: AsyncSession, project_id: int
) -> dict[int, list[ScenarioResult]]:
    """{scenario_id: list[ScenarioResult]} для всех сценариев проекта."""
    scenarios = (
        await session.scalars(
            select(Scenario).where(Scenario.project_id == project_id)
        )
    ).all()
    out: dict[int, list[ScenarioResult]] = {}
    for sc in scenarios:
        results = (
            await session.scalars(
                select(ScenarioResult).where(
                    ScenarioResult.scenario_id == sc.id
                )
            )
        ).all()
        out[sc.id] = list(results)
    return out


# ============================================================
# Sheet builders
# ============================================================


def _build_inputs_sheet(
    wb: Workbook,
    project: Project,
    inflation_profile: RefInflation | None,
    skus_with_bom: list[tuple[ProjectSKU, list[BOMItem]]],
    psk_channels: list[ProjectSKUChannel],
    financial_plan: list[ProjectFinancialPlan],
    period_by_id: dict[int, Period],
) -> None:
    """Лист «Вводные»: параметры проекта + SKU + каналы + financial plan."""
    ws = wb.create_sheet("Вводные")

    # === Project parameters ===
    _set_section(ws, 1, 1, "ПАРАМЕТРЫ ПРОЕКТА")
    params = [
        ("Название", project.name),
        ("Дата старта", project.start_date.isoformat()),
        ("Горизонт, лет", project.horizon_years),
        ("WACC", float(project.wacc)),
        ("Налог на прибыль", float(project.tax_rate)),
        ("Working Capital rate", float(project.wc_rate)),
        ("VAT rate", float(project.vat_rate)),
        ("Валюта", project.currency),
        (
            "Профиль инфляции",
            inflation_profile.profile_name if inflation_profile else "—",
        ),
    ]
    for i, (label, value) in enumerate(params, start=2):
        ws.cell(row=i, column=1, value=label).font = LABEL_FONT
        ws.cell(row=i, column=2, value=value)

    # === SKU table ===
    skus_start = len(params) + 4  # пустая строка + section header
    _set_section(ws, skus_start, 1, "SKU И BOM")
    # Q6 (2026-05-15): CA&M и Marketing переехали на ProjectSKUChannel
    # — в этом листе оставляем только SKU-уровневые поля, ставки
    # выводятся в листе "КАНАЛЫ × SKU" ниже.
    headers = [
        "SKU",
        "Бренд",
        "Объём, л",
        "Production rate",
        "BOM total ₽/unit",
    ]
    _set_header(ws, skus_start + 1, headers)
    row = skus_start + 2
    for psk, bom in skus_with_bom:
        bom_total = sum(
            float(b.quantity_per_unit * b.price_per_unit * (1 + b.loss_pct))
            for b in bom
        )
        ws.cell(row=row, column=1, value=psk.sku.name)
        ws.cell(row=row, column=2, value=psk.sku.brand)
        ws.cell(
            row=row,
            column=3,
            value=float(psk.sku.volume_l) if psk.sku.volume_l else None,
        )
        ws.cell(row=row, column=4, value=float(psk.production_cost_rate))
        ws.cell(row=row, column=5, value=round(bom_total, 4))
        row += 1

    # === Channels table ===
    chan_start = row + 2
    _set_section(ws, chan_start, 1, "КАНАЛЫ × SKU")
    headers = [
        "SKU ID",
        "Канал",
        "Launch Y/M",
        "ND target",
        "Off-take target",
        "Channel margin",
        "Promo discount",
        "Promo share",
        "Shelf price M1",
        "Logistics ₽/кг M1",
        "CA&M rate",       # Q6 (2026-05-15)
        "Marketing rate",  # Q6 (2026-05-15)
    ]
    _set_header(ws, chan_start + 1, headers)
    row = chan_start + 2
    for psc in psk_channels:
        ws.cell(row=row, column=1, value=psc.project_sku_id)
        ws.cell(row=row, column=2, value=psc.channel.code)
        ws.cell(
            row=row, column=3, value=f"Y{psc.launch_year}/M{psc.launch_month:02d}"
        )
        ws.cell(row=row, column=4, value=float(psc.nd_target))
        ws.cell(row=row, column=5, value=float(psc.offtake_target))
        ws.cell(row=row, column=6, value=float(psc.channel_margin))
        ws.cell(row=row, column=7, value=float(psc.promo_discount))
        ws.cell(row=row, column=8, value=float(psc.promo_share))
        ws.cell(row=row, column=9, value=float(psc.shelf_price_reg))
        ws.cell(row=row, column=10, value=float(psc.logistics_cost_per_kg))
        ws.cell(row=row, column=11, value=float(psc.ca_m_rate))
        ws.cell(row=row, column=12, value=float(psc.marketing_rate))
        row += 1

    # === Financial plan table (CAPEX/OPEX) ===
    if financial_plan:
        fp_start = row + 2
        _set_section(ws, fp_start, 1, "PROJECT FINANCIAL PLAN (CAPEX/OPEX)")
        _set_header(ws, fp_start + 1, ["Период", "Год", "CAPEX, ₽", "OPEX, ₽"])
        row = fp_start + 2
        for plan in sorted(
            financial_plan,
            key=lambda p: period_by_id[p.period_id].period_number,
        ):
            period = period_by_id[plan.period_id]
            label = (
                f"M{period.period_number}"
                if period.type == PeriodType.MONTHLY
                else f"Y{period.model_year}"
            )
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=period.model_year)
            ws.cell(row=row, column=3, value=float(plan.capex))
            ws.cell(row=row, column=4, value=float(plan.opex))
            row += 1

    _autosize_columns(ws)


def _period_label(p: Period) -> str:
    """Удобная метка периода для заголовка колонки."""
    if p.type == PeriodType.MONTHLY:
        return f"M{p.period_number}"
    return f"Y{p.model_year}"


def _build_pnl_sheet(
    wb: Workbook,
    sorted_periods: list[Period],
    base_aggregate: Any,
) -> None:
    """Лист «PnL по периодам» Base сценария.

    Базовый сценарий — основной для отображения в Excel. Per-period
    показатели берутся из агрегатного `PipelineContext` после
    `run_project_pipeline`.

    Строки = метрики, столбцы = периоды M1..M36 + Y4..Y10.
    """
    ws = wb.create_sheet("PnL по периодам")

    # Header: первая колонка = "Метрика", остальные = периоды
    headers = ["Метрика"] + [_period_label(p) for p in sorted_periods]
    _set_header(ws, 1, headers)

    # Метрики из PipelineContext (per-period)
    metric_rows: list[tuple[str, list[float]]] = [
        ("Volume Units", base_aggregate.volume_units),
        ("Volume Liters", base_aggregate.volume_liters),
        ("Net Revenue, ₽", base_aggregate.net_revenue),
        ("COGS Material, ₽", base_aggregate.cogs_material),
        ("COGS Production, ₽", base_aggregate.cogs_production),
        ("COGS Total, ₽", base_aggregate.cogs_total),
        ("Gross Profit, ₽", base_aggregate.gross_profit),
        ("Logistics Cost, ₽", base_aggregate.logistics_cost),
        ("Contribution, ₽", base_aggregate.contribution),
        ("CA&M Cost, ₽", base_aggregate.ca_m_cost),
        ("Marketing Cost, ₽", base_aggregate.marketing_cost),
        ("EBITDA, ₽", base_aggregate.ebitda),
        ("Working Capital, ₽", base_aggregate.working_capital),
        ("ΔWC, ₽", base_aggregate.delta_working_capital),
        ("Tax, ₽", base_aggregate.tax),
        ("OCF, ₽", base_aggregate.operating_cash_flow),
        ("ICF, ₽", base_aggregate.investing_cash_flow),
        ("FCF, ₽", base_aggregate.free_cash_flow),
    ]
    for i, (label, values) in enumerate(metric_rows, start=2):
        ws.cell(row=i, column=1, value=label).font = LABEL_FONT
        for j, v in enumerate(values, start=2):
            ws.cell(row=i, column=j, value=round(v, 2) if v is not None else None)

    # Annual section (Y1..Y10)
    if base_aggregate.annual_free_cash_flow:
        annual_start = len(metric_rows) + 4
        _set_section(ws, annual_start, 1, "ГОДОВЫЕ АГРЕГАТЫ (D-22)")
        annual_headers = ["Метрика"] + [
            f"Y{y}" for y in range(1, len(base_aggregate.annual_free_cash_flow) + 1)
        ]
        _set_header(ws, annual_start + 1, annual_headers)
        annual_rows: list[tuple[str, list[float]]] = [
            ("Annual NR, ₽", base_aggregate.annual_net_revenue),
            ("Annual CM, ₽", base_aggregate.annual_contribution),
            ("Annual FCF, ₽", base_aggregate.annual_free_cash_flow),
            ("Discounted FCF, ₽", base_aggregate.annual_discounted_cash_flow),
            ("Cumulative FCF, ₽", base_aggregate.cumulative_fcf),
            ("Cumulative DCF, ₽", base_aggregate.cumulative_dcf),
        ]
        for i, (label, values) in enumerate(annual_rows, start=annual_start + 2):
            ws.cell(row=i, column=1, value=label).font = LABEL_FONT
            for j, v in enumerate(values, start=2):
                ws.cell(
                    row=i,
                    column=j,
                    value=round(v, 2) if v is not None else None,
                )

    _autosize_columns(ws)


def _build_kpi_sheet(
    wb: Workbook,
    scenarios: list[Scenario],
    results_by_scenario: dict[int, list[ScenarioResult]],
) -> None:
    """Лист «KPI»: сводная таблица по 3 сценариям × 3 scope.

    Если `ScenarioResult` отсутствует для какого-то (сценарий × scope) —
    в ячейке "—". Это значит расчёт ещё не выполнен (POST /recalculate
    нужно нажать перед экспортом).
    """
    ws = wb.create_sheet("KPI")

    headers = [
        "Сценарий",
        "Scope",
        "NPV, ₽",
        "IRR",
        "ROI",
        "Payback simple",
        "Payback discounted",
        "Contribution Margin",
        "EBITDA Margin",
        "Go/No-Go",
    ]
    _set_header(ws, 1, headers)

    # Сортируем сценарии: Base / Conservative / Aggressive
    scenario_order = {
        ScenarioType.BASE: 0,
        ScenarioType.CONSERVATIVE: 1,
        ScenarioType.AGGRESSIVE: 2,
    }
    sorted_scenarios = sorted(
        scenarios, key=lambda s: scenario_order.get(s.type, 99)
    )

    scope_order = [PeriodScope.Y1Y3, PeriodScope.Y1Y5, PeriodScope.Y1Y10]
    scope_label = {
        PeriodScope.Y1Y3: "Y1-Y3",
        PeriodScope.Y1Y5: "Y1-Y5",
        PeriodScope.Y1Y10: "Y1-Y10",
    }

    row = 2
    for sc in sorted_scenarios:
        results = results_by_scenario.get(sc.id, [])
        results_by_scope = {r.period_scope: r for r in results}
        for scope in scope_order:
            r = results_by_scope.get(scope)
            ws.cell(row=row, column=1, value=sc.type.value)
            ws.cell(row=row, column=2, value=scope_label[scope])
            if r is None:
                # Расчёт не выполнен — заполняем "—"
                for col in range(3, 11):
                    ws.cell(row=row, column=col, value="—")
            else:
                ws.cell(
                    row=row,
                    column=3,
                    value=float(r.npv) if r.npv is not None else "—",
                )
                ws.cell(
                    row=row,
                    column=4,
                    value=float(r.irr) if r.irr is not None else "—",
                )
                ws.cell(
                    row=row,
                    column=5,
                    value=float(r.roi) if r.roi is not None else "—",
                )
                ws.cell(
                    row=row,
                    column=6,
                    value=float(r.payback_simple)
                    if r.payback_simple is not None
                    else "НЕ ОКУПАЕТСЯ",
                )
                ws.cell(
                    row=row,
                    column=7,
                    value=float(r.payback_discounted)
                    if r.payback_discounted is not None
                    else "НЕ ОКУПАЕТСЯ",
                )
                ws.cell(
                    row=row,
                    column=8,
                    value=float(r.contribution_margin)
                    if r.contribution_margin is not None
                    else "—",
                )
                ws.cell(
                    row=row,
                    column=9,
                    value=float(r.ebitda_margin)
                    if r.ebitda_margin is not None
                    else "—",
                )
                ws.cell(
                    row=row,
                    column=10,
                    value="GO" if r.go_no_go else "NO-GO",
                )
            row += 1

    _autosize_columns(ws)


# ============================================================
# C #15: P&L Pivot sheet
# ============================================================


def _build_pnl_pivot_sheet(
    wb: Workbook,
    sorted_periods: list[Period],
    line_pairs: list[tuple[PipelineInput, PipelineContext]],
    psc_by_id: dict[int, ProjectSKUChannel],
    psk_by_id: dict[int, ProjectSKU],
) -> None:
    """C #15: Лист «P&L Pivot» — per-line breakdown для сводных таблиц.

    Каждая строка = (SKU × Channel × Period). Пользователь строит
    сводные таблицы в Excel нативно.

    Args:
        sorted_periods: 43 периода в порядке period_number (M1..M36, Y4..Y10).
        line_pairs: список (PipelineInput, PipelineContext) — по одной паре
            на каждую (PSK × Channel) линию, уже прогнанную через s01..s09.
        psc_by_id: {psc_id: ProjectSKUChannel} с подгруженным .channel.
        psk_by_id: {psk_id: ProjectSKU} с подгруженным .sku.
    """
    ws = wb.create_sheet("P&L Pivot")

    # Header row
    _set_header(ws, 1, PNL_PIVOT_HEADERS)

    row_idx = 2

    def _val_at(series: list[float], t: int) -> float | None:
        """Return rounded value at index t, or None if series is empty/short."""
        if not series or t >= len(series):
            return None
        return round(series[t], 2)

    for inp, ctx in line_pairs:
        psc = psc_by_id.get(inp.project_sku_channel_id)
        if psc is None:
            # Агрегатный инпут (project_sku_channel_id == 0) — пропускаем.
            continue

        psk = psk_by_id.get(psc.project_sku_id)
        if psk is None:
            continue

        sku = psk.sku
        channel = psc.channel

        # Quarter derived from month_num (1..12 → Q1..Q4)
        for period_idx, period in enumerate(sorted_periods):
            if period_idx >= inp.period_count:
                break

            month_num = period.month_num  # int | None
            quarter: int | None = None
            if month_num is not None:
                quarter = (month_num - 1) // 3 + 1

            period_label = _period_label(period)
            period_type = period.type.value  # "monthly" or "annual"
            t = period_idx

            row_data = [
                sku.brand,
                sku.name,
                sku.format,
                float(sku.volume_l) if sku.volume_l is not None else None,
                sku.unit_of_measure,
                channel.code,
                channel.name,
                channel.channel_group,
                channel.source_type,
                period_label,
                period_type,
                period.model_year,
                month_num,
                quarter,
                _val_at(ctx.volume_units, t),
                _val_at(ctx.volume_liters, t),
                _val_at(ctx.net_revenue, t),
                _val_at(ctx.cogs_material, t),
                _val_at(ctx.cogs_production, t),
                _val_at(ctx.cogs_total, t),
                _val_at(ctx.gross_profit, t),
                _val_at(ctx.logistics_cost, t),
                _val_at(ctx.contribution, t),
                _val_at(ctx.ca_m_cost, t),
                _val_at(ctx.marketing_cost, t),
                _val_at(ctx.ebitda, t),
            ]

            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

            row_idx += 1

    # Auto-size: use existing helper for consistent look
    _autosize_columns(ws, max_width=25)


# ============================================================
# Public entry point
# ============================================================


class ProjectNotFoundForExport(Exception):
    """project_id не существует или soft-deleted."""


async def generate_project_xlsx(
    session: AsyncSession,
    project_id: int,
) -> bytes:
    """Главная функция — генерирует XLSX для проекта и возвращает bytes.

    Шаги:
    1. Загружает project + inflation profile + SKU/BOM + каналы +
       financial plan + scenarios + scenario results
    2. Запускает Base пайплайн in-memory чтобы получить per-period
       детали для PnL листа (ScenarioResult хранит только KPI agg)
    3. Строит 3 листа в openpyxl Workbook
    4. Сериализует в BytesIO → bytes

    Raises:
        ProjectNotFoundForExport: если project_id не существует.
    """
    project = await _load_project_full(session, project_id)
    if project is None:
        raise ProjectNotFoundForExport(f"Project {project_id} not found")

    # Inflation profile (опционально)
    inflation_profile: RefInflation | None = None
    if project.inflation_profile_id is not None:
        inflation_profile = await session.get(
            RefInflation, project.inflation_profile_id
        )

    # SKU + BOM
    skus_with_bom = await _load_skus_with_bom(session, project_id)

    # Channels
    psk_channels = await _load_psk_channels(session, project_id)

    # Financial plan
    fp_rows = (
        await session.scalars(
            select(ProjectFinancialPlan).where(
                ProjectFinancialPlan.project_id == project_id
            )
        )
    ).all()

    # Periods catalog (для labels и financial plan join)
    sorted_periods, period_by_id = await _load_period_catalog(session)

    # Scenarios + results
    scenarios = (
        await session.scalars(
            select(Scenario).where(Scenario.project_id == project_id)
        )
    ).all()
    results_by_scenario = await _load_scenario_results(session, project_id)

    # Base сценарий: запускаем pipeline чтобы получить per-period детали
    # для PnL листа. Если нет PSC — pipeline кинет NoLinesError, ловим.
    base_scenario = next(
        (s for s in scenarios if s.type == ScenarioType.BASE), None
    )
    base_aggregate = None
    # C #15: per-line (PipelineInput, PipelineContext) pairs for pivot sheet.
    line_pairs: list[tuple[PipelineInput, PipelineContext]] = []
    if base_scenario is not None and skus_with_bom and psk_channels:
        try:
            line_inputs = await build_line_inputs(
                session, project_id, base_scenario.id
            )
            capex, opex = await _load_project_financial_plan(
                session, project_id, sorted_periods
            )
            # C #15: run per-line pipeline explicitly to capture individual
            # PipelineContext objects needed for the pivot sheet.
            # Then aggregate + run s10..s12 manually (same as run_project_pipeline).
            line_contexts = [run_line_pipeline(inp) for inp in line_inputs]
            line_pairs = list(zip(line_inputs, line_contexts))

            from app.engine.aggregator import aggregate_lines

            agg = aggregate_lines(
                line_contexts, project_capex=capex, project_opex=opex
            )
            # Apply project-level OPEX/CAPEX adjustments (mirrors run_project_pipeline)
            n = agg.input.period_count
            if opex:
                for t in range(n):
                    agg.contribution[t] -= opex[t]
                    agg.operating_cash_flow[t] -= opex[t]
                    agg.free_cash_flow[t] -= opex[t]
            if capex:
                for t in range(n):
                    agg.investing_cash_flow[t] -= capex[t]
                    agg.free_cash_flow[t] -= capex[t]

            from app.engine.steps import s10_discount, s11_kpi, s12_gonogo

            s10_discount.step(agg)
            s11_kpi.step(agg)
            s12_gonogo.step(agg)
            base_aggregate = agg
        except Exception:  # noqa: BLE001
            # Если pipeline падает (например, нет данных) — экспорт всё
            # равно генерирует листы Вводные/KPI, PnL пропускается.
            base_aggregate = None
            line_pairs = []

    # Build workbook
    wb = Workbook()
    # Удаляем default лист "Sheet"
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)

    _build_inputs_sheet(
        wb,
        project,
        inflation_profile,
        skus_with_bom,
        psk_channels,
        list(fp_rows),
        period_by_id,
    )

    if base_aggregate is not None:
        _build_pnl_sheet(wb, sorted_periods, base_aggregate)
    else:
        # Placeholder лист с пометкой
        ws = wb.create_sheet("PnL по периодам")
        ws.cell(
            row=1,
            column=1,
            value="Расчёт не выполнен. Запустите POST /api/projects/{id}/recalculate.",
        )

    _build_kpi_sheet(wb, list(scenarios), results_by_scenario)

    # C #15: P&L Pivot sheet — per-line breakdown for pivot tables in Excel.
    # Build lookup maps from already-loaded ORM objects (no extra DB queries).
    psc_by_id: dict[int, ProjectSKUChannel] = {psc.id: psc for psc in psk_channels}
    psk_by_id: dict[int, ProjectSKU] = {psk.id: psk for psk, _ in skus_with_bom}
    if line_pairs:
        _build_pnl_pivot_sheet(
            wb,
            sorted_periods,
            line_pairs,
            psc_by_id,
            psk_by_id,
        )
    else:
        # No pipeline data — placeholder sheet so the tab always exists.
        ws_pivot = wb.create_sheet("P&L Pivot")
        ws_pivot.cell(
            row=1,
            column=1,
            value="Расчёт не выполнен. Запустите POST /api/projects/{id}/recalculate.",
        )

    # Сериализация
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
