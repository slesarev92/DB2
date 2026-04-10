"""Service для импорта фактических данных (actual) из Excel.

B-02: пользователь загружает .xlsx с фактическими показателями
(ND, offtake, shelf_price) по SKU × Channel × Period.

Формат Excel (один sheet):
  | Period | SKU   | Channel | nd    | offtake | shelf_price |
  | M1     | GORJI | HM      | 0.15  | 45      | 89.50       |
  | M2     | GORJI | HM      | 0.20  | 48      | 89.50       |
  | ...    |       |         |       |         |             |

- Period — label вида M1..M36, Y4..Y10 (case-insensitive)
- SKU — sku.name (точное совпадение, case-insensitive)
- Channel — channel.code (точное совпадение, case-insensitive)
- nd, offtake, shelf_price — числовые значения (опциональные: пустые ячейки пропускаются)

Каждая валидная строка создаёт PeriodValue с source_type=ACTUAL.
При повторном импорте version_id инкрементируется (append-only).
"""
from __future__ import annotations

import io
from dataclasses import dataclass
from decimal import Decimal
from typing import Any, BinaryIO

from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models import (
    Channel,
    Period,
    PeriodValue,
    ProjectSKU,
    ProjectSKUChannel,
    SKU,
    Scenario,
    SourceType,
)


# ============================================================
# Exceptions
# ============================================================


class ActualImportError(Exception):
    """Ошибка импорта actual-данных."""


class EmptyFileError(ActualImportError):
    """Файл пустой или нет данных."""


class InvalidFormatError(ActualImportError):
    """Отсутствуют обязательные колонки."""


# ============================================================
# Result dataclass
# ============================================================


@dataclass
class ImportResult:
    """Результат импорта."""

    imported: int
    skipped: int
    errors: list[str]


# ============================================================
# Constants
# ============================================================

REQUIRED_COLUMNS = {"period", "sku", "channel"}
VALUE_COLUMNS = {"nd", "offtake", "shelf_price"}
PERIOD_LABEL_MAP: dict[str, int] = {}  # populated lazily


def _build_period_label_to_number() -> dict[str, int]:
    """M1..M36 → 1..36, Y4..Y10 → 37..43."""
    mapping: dict[str, int] = {}
    for i in range(1, 37):
        mapping[f"m{i}"] = i
    for y in range(4, 11):
        mapping[f"y{y}"] = 36 + (y - 3)  # Y4=37, Y5=38, ..., Y10=43
    return mapping


PERIOD_LABEL_TO_NUMBER = _build_period_label_to_number()


# ============================================================
# Excel parsing
# ============================================================


def _parse_excel(fileobj: BinaryIO) -> list[dict[str, Any]]:
    """Парсит первый лист Excel в список словарей.

    Возвращает строки с ключами в lowercase. Пропускает пустые строки.
    """
    wb = load_workbook(fileobj, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise EmptyFileError("Excel файл не содержит листов")

    rows_iter = ws.iter_rows(values_only=True)

    # Header row
    header_row = next(rows_iter, None)
    if header_row is None:
        raise EmptyFileError("Excel файл пуст")

    headers = [str(h).strip().lower() if h is not None else "" for h in header_row]

    # Validate required columns
    missing = REQUIRED_COLUMNS - set(headers)
    if missing:
        raise InvalidFormatError(
            f"Отсутствуют обязательные колонки: {', '.join(sorted(missing))}. "
            f"Найдены: {', '.join(h for h in headers if h)}"
        )

    # Parse data rows
    result: list[dict[str, Any]] = []
    for row_values in rows_iter:
        row = dict(zip(headers, row_values, strict=False))
        # Skip fully empty rows
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in row_values):
            continue
        result.append(row)

    wb.close()
    return result


# ============================================================
# Import logic
# ============================================================


async def import_actual_data(
    session: AsyncSession,
    project_id: int,
    scenario_id: int,
    fileobj: BinaryIO,
) -> ImportResult:
    """Импортирует actual-данные из Excel в PeriodValue.

    Алгоритм:
    1. Парсим Excel → список строк
    2. Загружаем справочники: periods, project SKU/channels
    3. Для каждой строки: валидируем, ищём psk_channel_id, создаём PeriodValue
    4. Возвращаем результат (imported, skipped, errors)
    """
    rows = _parse_excel(fileobj)
    if not rows:
        raise EmptyFileError("Excel файл не содержит строк данных")

    # Load period catalog
    periods = (
        await session.scalars(select(Period).order_by(Period.period_number))
    ).all()
    period_num_to_id = {p.period_number: p.id for p in periods}

    # Load project SKUs with their SKU names
    psk_rows = (
        await session.execute(
            select(ProjectSKU, SKU)
            .join(SKU, SKU.id == ProjectSKU.sku_id)
            .where(ProjectSKU.project_id == project_id)
            .where(ProjectSKU.include.is_(True))
        )
    ).all()
    # sku_name (lower) → project_sku_id
    sku_name_to_psk: dict[str, int] = {}
    for psk, sku in psk_rows:
        sku_name_to_psk[sku.name.lower()] = psk.id

    # Load channels for this project's PSKs
    psk_ids = list(sku_name_to_psk.values())
    if not psk_ids:
        raise ActualImportError("В проекте нет SKU")

    psc_rows = (
        await session.execute(
            select(ProjectSKUChannel, Channel)
            .join(Channel, Channel.id == ProjectSKUChannel.channel_id)
            .where(ProjectSKUChannel.project_sku_id.in_(psk_ids))
                    )
    ).all()
    # (psk_id, channel_code_lower) → psc_id
    psk_channel_map: dict[tuple[int, str], int] = {}
    for psc, ch in psc_rows:
        psk_channel_map[(psc.project_sku_id, ch.code.lower())] = psc.id

    # Validate scenario belongs to project
    scenario = await session.get(Scenario, scenario_id)
    if scenario is None or scenario.project_id != project_id:
        raise ActualImportError(
            f"Сценарий {scenario_id} не найден или не принадлежит проекту"
        )

    # Pre-load existing actual max version_ids for batch efficiency
    existing_versions = await _load_existing_actual_versions(
        session, [v for v in psk_channel_map.values()], scenario_id
    )

    imported = 0
    skipped = 0
    errors: list[str] = []

    for row_idx, row in enumerate(rows, start=2):  # Excel row 2+ (1=header)
        period_label = str(row.get("period", "")).strip().lower()
        sku_name = str(row.get("sku", "")).strip().lower()
        channel_code = str(row.get("channel", "")).strip().lower()

        # Validate period
        period_number = PERIOD_LABEL_TO_NUMBER.get(period_label)
        if period_number is None:
            errors.append(f"Строка {row_idx}: неизвестный период '{row.get('period')}'")
            skipped += 1
            continue

        period_id = period_num_to_id.get(period_number)
        if period_id is None:
            errors.append(f"Строка {row_idx}: период {period_label} не найден в БД")
            skipped += 1
            continue

        # Validate SKU
        psk_id = sku_name_to_psk.get(sku_name)
        if psk_id is None:
            errors.append(f"Строка {row_idx}: SKU '{row.get('sku')}' не найден в проекте")
            skipped += 1
            continue

        # Validate Channel
        psc_id = psk_channel_map.get((psk_id, channel_code))
        if psc_id is None:
            errors.append(
                f"Строка {row_idx}: канал '{row.get('channel')}' "
                f"не найден для SKU '{row.get('sku')}'"
            )
            skipped += 1
            continue

        # Build values dict (only non-empty numeric fields)
        values: dict[str, Any] = {}
        for col in VALUE_COLUMNS:
            raw = row.get(col)
            if raw is not None and raw != "":
                try:
                    values[col] = float(raw)
                except (ValueError, TypeError):
                    errors.append(
                        f"Строка {row_idx}: '{col}' = '{raw}' — не число, пропущено"
                    )

        if not values:
            skipped += 1
            continue

        # Determine version_id (append-only)
        version_key = (psc_id, period_id)
        current_max = existing_versions.get(version_key, 0)
        new_version = current_max + 1
        existing_versions[version_key] = new_version

        session.add(
            PeriodValue(
                psk_channel_id=psc_id,
                scenario_id=scenario_id,
                period_id=period_id,
                values=values,
                source_type=SourceType.ACTUAL,
                version_id=new_version,
                is_overridden=False,
            )
        )
        imported += 1

    await session.flush()
    return ImportResult(imported=imported, skipped=skipped, errors=errors)


async def _load_existing_actual_versions(
    session: AsyncSession,
    psc_ids: list[int],
    scenario_id: int,
) -> dict[tuple[int, int], int]:
    """Загружает MAX(version_id) для existing actual PeriodValues.

    Возвращает {(psc_id, period_id): max_version_id}.
    """
    if not psc_ids:
        return {}

    stmt = (
        select(
            PeriodValue.psk_channel_id,
            PeriodValue.period_id,
            func.max(PeriodValue.version_id),
        )
        .where(PeriodValue.psk_channel_id.in_(psc_ids))
        .where(PeriodValue.scenario_id == scenario_id)
        .where(PeriodValue.source_type == SourceType.ACTUAL)
        .group_by(PeriodValue.psk_channel_id, PeriodValue.period_id)
    )

    rows = (await session.execute(stmt)).all()
    return {(r[0], r[1]): r[2] for r in rows}


# ============================================================
# Template generation
# ============================================================


async def generate_template(
    session: AsyncSession,
    project_id: int,
) -> bytes:
    """Генерирует пустой Excel-шаблон для импорта actual-данных.

    Шаблон содержит:
    - Заголовки: Period, SKU, Channel, nd, offtake, shelf_price
    - Строки для каждой комбинации (Period × SKU × Channel) проекта
    """
    # Load periods
    periods = (
        await session.scalars(select(Period).order_by(Period.period_number))
    ).all()

    # Load project SKUs
    psk_rows = (
        await session.execute(
            select(ProjectSKU, SKU)
            .join(SKU, SKU.id == ProjectSKU.sku_id)
            .where(ProjectSKU.project_id == project_id)
            .where(ProjectSKU.include.is_(True))
        )
    ).all()

    psk_data: list[tuple[int, str]] = []  # (psk_id, sku_name)
    for psk, sku in psk_rows:
        psk_data.append((psk.id, sku.name))

    # Load channels per PSK
    psk_ids = [p[0] for p in psk_data]
    psc_rows = (
        await session.execute(
            select(ProjectSKUChannel, Channel)
            .join(Channel, Channel.id == ProjectSKUChannel.channel_id)
            .where(ProjectSKUChannel.project_sku_id.in_(psk_ids))
                    )
    ).all() if psk_ids else []

    # psk_id → list of channel codes
    psk_channels: dict[int, list[str]] = {}
    for psc, ch in psc_rows:
        psk_channels.setdefault(psc.project_sku_id, []).append(ch.code)

    # Generate workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Actual Data"

    # Header
    ws.append(["Period", "SKU", "Channel", "nd", "offtake", "shelf_price"])

    # Period labels
    period_labels: list[str] = []
    for p in periods:
        if p.period_number <= 36:
            period_labels.append(f"M{p.period_number}")
        else:
            period_labels.append(f"Y{p.model_year}")

    # One row per (Period × SKU × Channel)
    for psk_id, sku_name in psk_data:
        channels = psk_channels.get(psk_id, [])
        for ch_code in sorted(channels):
            for label in period_labels:
                ws.append([label, sku_name, ch_code, None, None, None])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
