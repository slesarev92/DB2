"""Тесты B-02: импорт фактических данных из Excel.

POST /api/projects/{id}/actual-import?scenario_id
GET  /api/projects/{id}/actual-import/template
"""
import io

from httpx import AsyncClient
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import (
    Channel,
    Period,
    PeriodValue,
    Scenario,
    ScenarioType,
    SourceType,
)


# ============================================================
# Helpers
# ============================================================


SKU_BODY = {"brand": "Gorji", "name": "Test Import SKU"}
PROJECT_BODY = {"name": "Actual import test", "start_date": "2025-01-01"}


async def _setup(
    auth_client: AsyncClient, db_session: AsyncSession
) -> tuple[int, int]:
    """Создаёт project + sku + psk + psc (без auto_fill_predict).

    Возвращает (project_id, base_scenario_id).
    """
    from app.schemas.project_sku_channel import ProjectSKUChannelCreate
    from app.services.project_sku_channel_service import create_psk_channel

    project_id = (
        await auth_client.post("/api/projects", json=PROJECT_BODY)
    ).json()["id"]
    sku_id = (
        await auth_client.post("/api/skus", json=SKU_BODY)
    ).json()["id"]
    psk_id = (
        await auth_client.post(
            f"/api/projects/{project_id}/skus", json={"sku_id": sku_id}
        )
    ).json()["id"]

    hm = await db_session.scalar(select(Channel).where(Channel.code == "HM"))
    assert hm is not None

    await create_psk_channel(
        db_session,
        psk_id,
        ProjectSKUChannelCreate(channel_id=hm.id),
        auto_fill_predict=False,
    )

    base_scenario = await db_session.scalar(
        select(Scenario).where(
            Scenario.project_id == project_id,
            Scenario.type == ScenarioType.BASE,
        )
    )
    assert base_scenario is not None

    return project_id, base_scenario.id


def _make_xlsx(rows: list[list]) -> io.BytesIO:
    """Создаёт xlsx в памяти. Первая строка — заголовок."""
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ============================================================
# POST /actual-import — happy path
# ============================================================


async def test_import_actual_basic(
    auth_client: AsyncClient, db_session: AsyncSession
) -> None:
    """Загрузка Excel с 2 строками → 2 actual PeriodValues."""
    project_id, scenario_id = await _setup(auth_client, db_session)

    xlsx = _make_xlsx([
        ["Period", "SKU", "Channel", "nd", "offtake", "shelf_price"],
        ["M1", "Test Import SKU", "HM", 0.15, 45, 89.50],
        ["M2", "Test Import SKU", "HM", 0.20, 48, 92.00],
    ])

    resp = await auth_client.post(
        f"/api/projects/{project_id}/actual-import",
        params={"scenario_id": scenario_id},
        files={"file": ("actual.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["imported"] == 2
    assert data["skipped"] == 0
    assert data["errors"] == []

    # Verify PeriodValues in DB
    pvs = (
        await db_session.scalars(
            select(PeriodValue).where(
                PeriodValue.source_type == SourceType.ACTUAL
            )
        )
    ).all()
    assert len(list(pvs)) == 2
    pv0 = pvs[0]
    assert pv0.version_id == 1
    assert pv0.source_type == SourceType.ACTUAL


# ============================================================
# POST /actual-import — repeat upload increments version
# ============================================================


async def test_import_actual_repeat_increments_version(
    auth_client: AsyncClient, db_session: AsyncSession
) -> None:
    """Повторный импорт тех же периодов → version_id=2."""
    project_id, scenario_id = await _setup(auth_client, db_session)

    xlsx1 = _make_xlsx([
        ["Period", "SKU", "Channel", "nd", "offtake", "shelf_price"],
        ["M1", "Test Import SKU", "HM", 0.10, 30, 80.0],
    ])
    await auth_client.post(
        f"/api/projects/{project_id}/actual-import",
        params={"scenario_id": scenario_id},
        files={"file": ("actual.xlsx", xlsx1, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    xlsx2 = _make_xlsx([
        ["Period", "SKU", "Channel", "nd", "offtake", "shelf_price"],
        ["M1", "Test Import SKU", "HM", 0.15, 45, 89.50],
    ])
    resp = await auth_client.post(
        f"/api/projects/{project_id}/actual-import",
        params={"scenario_id": scenario_id},
        files={"file": ("actual.xlsx", xlsx2, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200
    assert resp.json()["imported"] == 1

    # Both versions exist (append-only)
    pvs = (
        await db_session.scalars(
            select(PeriodValue).where(
                PeriodValue.source_type == SourceType.ACTUAL
            ).order_by(PeriodValue.version_id)
        )
    ).all()
    assert len(list(pvs)) == 2
    assert pvs[0].version_id == 1
    assert pvs[1].version_id == 2
    assert pvs[1].values["nd"] == 0.15


# ============================================================
# POST /actual-import — invalid SKU → error in response
# ============================================================


async def test_import_actual_invalid_sku_reports_error(
    auth_client: AsyncClient, db_session: AsyncSession
) -> None:
    project_id, scenario_id = await _setup(auth_client, db_session)

    xlsx = _make_xlsx([
        ["Period", "SKU", "Channel", "nd", "offtake", "shelf_price"],
        ["M1", "Nonexistent SKU", "HM", 0.15, 45, 89.50],
    ])
    resp = await auth_client.post(
        f"/api/projects/{project_id}/actual-import",
        params={"scenario_id": scenario_id},
        files={"file": ("actual.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["imported"] == 0
    assert data["skipped"] == 1
    assert len(data["errors"]) == 1
    assert "SKU" in data["errors"][0]


# ============================================================
# POST /actual-import — missing columns → 400
# ============================================================


async def test_import_actual_missing_columns(
    auth_client: AsyncClient, db_session: AsyncSession
) -> None:
    project_id, scenario_id = await _setup(auth_client, db_session)

    xlsx = _make_xlsx([
        ["Period", "nd", "offtake"],  # Missing SKU, Channel
        ["M1", 0.15, 45],
    ])
    resp = await auth_client.post(
        f"/api/projects/{project_id}/actual-import",
        params={"scenario_id": scenario_id},
        files={"file": ("actual.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 400
    assert "колонки" in resp.json()["detail"].lower()


# ============================================================
# POST /actual-import — empty file → 400
# ============================================================


async def test_import_actual_empty_file(
    auth_client: AsyncClient, db_session: AsyncSession
) -> None:
    project_id, scenario_id = await _setup(auth_client, db_session)

    xlsx = _make_xlsx([
        ["Period", "SKU", "Channel", "nd", "offtake", "shelf_price"],
        # No data rows
    ])
    resp = await auth_client.post(
        f"/api/projects/{project_id}/actual-import",
        params={"scenario_id": scenario_id},
        files={"file": ("actual.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 400


# ============================================================
# POST /actual-import — auth required
# ============================================================


async def test_import_actual_unauthorized(client: AsyncClient) -> None:
    resp = await client.post(
        "/api/projects/1/actual-import",
        params={"scenario_id": 1},
        files={"file": ("actual.xlsx", b"fake", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 401


# ============================================================
# GET /actual-import/template — happy path
# ============================================================


async def test_download_template(
    auth_client: AsyncClient, db_session: AsyncSession
) -> None:
    """Скачать шаблон → валидный xlsx с заголовками."""
    project_id, _ = await _setup(auth_client, db_session)

    resp = await auth_client.get(
        f"/api/projects/{project_id}/actual-import/template"
    )
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    assert "attachment" in resp.headers["content-disposition"]

    # Parse downloaded xlsx
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    assert "Period" in headers
    assert "SKU" in headers
    assert "Channel" in headers
    assert "nd" in headers
    # Should have rows (43 periods × 1 SKU × 1 channel = 43 data rows + 1 header)
    assert ws.max_row == 44


# ============================================================
# GET /actual-import/template — auth required
# ============================================================


async def test_download_template_unauthorized(client: AsyncClient) -> None:
    resp = await client.get("/api/projects/1/actual-import/template")
    assert resp.status_code == 401


# ============================================================
# POST /actual-import — partial values (only nd, no offtake/shelf_price)
# ============================================================


async def test_import_actual_partial_values(
    auth_client: AsyncClient, db_session: AsyncSession
) -> None:
    """Строка с заполненным только nd → импортируется, values содержит только nd."""
    project_id, scenario_id = await _setup(auth_client, db_session)

    xlsx = _make_xlsx([
        ["Period", "SKU", "Channel", "nd", "offtake", "shelf_price"],
        ["M5", "Test Import SKU", "HM", 0.30, None, None],
    ])
    resp = await auth_client.post(
        f"/api/projects/{project_id}/actual-import",
        params={"scenario_id": scenario_id},
        files={"file": ("actual.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200
    assert resp.json()["imported"] == 1

    pvs = (
        await db_session.scalars(
            select(PeriodValue).where(
                PeriodValue.source_type == SourceType.ACTUAL
            )
        )
    ).all()
    assert len(list(pvs)) == 1
    assert pvs[0].values == {"nd": 0.30}
