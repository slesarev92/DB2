"""Integration tests для XLSX экспорта (5.1).

Стратегия:
- Service-level: generate_project_xlsx → bytes, parse openpyxl, проверить
  3 листа + ключевые ячейки.
- Endpoint-level: GET /api/projects/{id}/export/xlsx → 200 + правильный
  Content-Type + filename + valid XLSX bytes.
"""
from __future__ import annotations

from io import BytesIO

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.export.excel_exporter import (
    ProjectNotFoundForExport,
    generate_project_xlsx,
)
from tests.api.test_calculation import _seed_minimal_project


XLSX_MIME = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


# ============================================================
# Service-level tests
# ============================================================


class TestGenerateProjectXlsx:
    async def test_returns_bytes(self, db_session: AsyncSession):
        """Service возвращает bytes — это минимальное assertion."""
        project_id, _, _, _ = await _seed_minimal_project(db_session)
        result = await generate_project_xlsx(db_session, project_id)
        assert isinstance(result, bytes)
        assert len(result) > 0

    async def test_has_four_sheets(self, db_session: AsyncSession):
        """XLSX содержит 4 листа: Вводные / PnL / KPI / P&L Pivot (C #15)."""
        project_id, _, _, _ = await _seed_minimal_project(db_session)
        result = await generate_project_xlsx(db_session, project_id)

        wb = load_workbook(BytesIO(result), data_only=False)
        assert "Вводные" in wb.sheetnames
        assert "PnL по периодам" in wb.sheetnames
        assert "KPI" in wb.sheetnames
        assert "P&L Pivot" in wb.sheetnames

    async def test_inputs_sheet_has_project_params(
        self, db_session: AsyncSession
    ):
        """Лист «Вводные» содержит параметры проекта в первых строках."""
        project_id, _, _, _ = await _seed_minimal_project(db_session)
        result = await generate_project_xlsx(db_session, project_id)

        wb = load_workbook(BytesIO(result), data_only=False)
        ws = wb["Вводные"]

        # Section header
        assert ws.cell(row=1, column=1).value == "ПАРАМЕТРЫ ПРОЕКТА"
        # Первый параметр — Название
        assert ws.cell(row=2, column=1).value == "Название"
        assert ws.cell(row=2, column=2).value == "Calc test project"

    async def test_inputs_sheet_has_sku_table(
        self, db_session: AsyncSession
    ):
        """Лист «Вводные» содержит таблицу SKU + BOM total."""
        project_id, _, _, _ = await _seed_minimal_project(db_session)
        result = await generate_project_xlsx(db_session, project_id)

        wb = load_workbook(BytesIO(result), data_only=False)
        ws = wb["Вводные"]

        # Найти section "SKU И BOM"
        sku_section_row = None
        for row in range(1, 30):
            if ws.cell(row=row, column=1).value == "SKU И BOM":
                sku_section_row = row
                break
        assert sku_section_row is not None

        # Через 2 строки — данные
        first_sku_row = sku_section_row + 2
        # SKU name из minimal_project = "Calc test SKU"
        assert ws.cell(row=first_sku_row, column=1).value == "Calc test SKU"
        # BOM total = 10.0 × 1 × (1+0) = 10
        # Q6 (2026-05-15): убрали CA&M/Marketing колонки с SKU-листа,
        # BOM total теперь column 5 (раньше column 7).
        assert ws.cell(row=first_sku_row, column=5).value == pytest.approx(
            10.0
        )

    async def test_inputs_sheet_has_channels_table(
        self, db_session: AsyncSession
    ):
        """Лист «Вводные» содержит таблицу каналов с launch + параметрами."""
        project_id, _, _, _ = await _seed_minimal_project(db_session)
        result = await generate_project_xlsx(db_session, project_id)

        wb = load_workbook(BytesIO(result), data_only=False)
        ws = wb["Вводные"]

        chan_section_row = None
        for row in range(1, 50):
            if ws.cell(row=row, column=1).value == "КАНАЛЫ × SKU":
                chan_section_row = row
                break
        assert chan_section_row is not None

        first_chan_row = chan_section_row + 2
        # _seed_minimal_project использует HM канал
        assert ws.cell(row=first_chan_row, column=2).value == "HM"
        # Launch по умолчанию Y1/M01
        assert ws.cell(row=first_chan_row, column=3).value == "Y1/M01"

    async def test_pnl_sheet_has_period_columns(
        self, db_session: AsyncSession
    ):
        """PnL лист имеет 43 period колонок (M1..M36 + Y4..Y10)."""
        project_id, _, _, _ = await _seed_minimal_project(db_session)
        result = await generate_project_xlsx(db_session, project_id)

        wb = load_workbook(BytesIO(result), data_only=False)
        ws = wb["PnL по периодам"]

        # Header row 1: 1 = "Метрика", 2..44 = периоды
        assert ws.cell(row=1, column=1).value == "Метрика"
        # Первый period column = M1
        assert ws.cell(row=1, column=2).value == "M1"
        # Последний monthly = M36 в колонке 37 (1+36)
        assert ws.cell(row=1, column=37).value == "M36"
        # Первый yearly = Y4 в колонке 38
        assert ws.cell(row=1, column=38).value == "Y4"
        # Последний = Y10 в колонке 44
        assert ws.cell(row=1, column=44).value == "Y10"

    async def test_pnl_sheet_has_metric_rows(
        self, db_session: AsyncSession
    ):
        """PnL содержит ключевые метрики (Volume, NR, GP, FCF и т.д.)."""
        project_id, _, _, _ = await _seed_minimal_project(db_session)
        result = await generate_project_xlsx(db_session, project_id)

        wb = load_workbook(BytesIO(result), data_only=False)
        ws = wb["PnL по периодам"]

        metric_labels = []
        for row in range(2, 25):
            label = ws.cell(row=row, column=1).value
            if label is not None:
                metric_labels.append(label)

        # Проверяем что ключевые метрики есть
        assert "Volume Units" in metric_labels
        assert "Net Revenue, ₽" in metric_labels
        assert "Gross Profit, ₽" in metric_labels
        assert "Contribution, ₽" in metric_labels
        assert "FCF, ₽" in metric_labels

    async def test_kpi_sheet_has_3x3_matrix(self, db_session: AsyncSession):
        """KPI лист содержит 3 сценария × 3 scope = 9 строк (после header)."""
        from app.services.calculation_service import calculate_all_scenarios

        project_id, _, _, _ = await _seed_minimal_project(db_session)
        # Запускаем calculate чтобы появились ScenarioResult
        await calculate_all_scenarios(db_session, project_id)
        await db_session.flush()

        result = await generate_project_xlsx(db_session, project_id)
        wb = load_workbook(BytesIO(result), data_only=False)
        ws = wb["KPI"]

        # Header row 1
        assert ws.cell(row=1, column=1).value == "Сценарий"
        assert ws.cell(row=1, column=2).value == "Scope"

        # 9 data rows (rows 2-10)
        scenarios_found = set()
        scopes_found = set()
        for row in range(2, 11):
            scenario_val = ws.cell(row=row, column=1).value
            scope_val = ws.cell(row=row, column=2).value
            if scenario_val:
                scenarios_found.add(scenario_val)
            if scope_val:
                scopes_found.add(scope_val)

        assert scenarios_found == {"base", "conservative", "aggressive"}
        assert scopes_found == {"Y1-Y3", "Y1-Y5", "Y1-Y10"}

    async def test_404_for_unknown_project(self, db_session: AsyncSession):
        with pytest.raises(ProjectNotFoundForExport):
            await generate_project_xlsx(db_session, 999999)

    # ----------------------------------------------------------
    # C #15: P&L Pivot sheet tests
    # ----------------------------------------------------------

    async def test_pnl_pivot_sheet_exists(self, db_session: AsyncSession):
        """C #15: XLSX содержит лист «P&L Pivot»."""
        project_id, _, _, _ = await _seed_minimal_project(db_session)
        result = await generate_project_xlsx(db_session, project_id)

        wb = load_workbook(BytesIO(result), data_only=False)
        assert "P&L Pivot" in wb.sheetnames

    async def test_pnl_pivot_sheet_has_per_line_rows(
        self, db_session: AsyncSession
    ):
        """C #15: лист «P&L Pivot» содержит header + per-line rows.

        1 PSC × 43 periods = 43 data rows + 1 header row = 44 строки минимум.
        26 колонок (PNL_PIVOT_HEADERS).
        """
        from app.export.excel_exporter import PNL_PIVOT_HEADERS

        project_id, _, _, _ = await _seed_minimal_project(db_session)
        result = await generate_project_xlsx(db_session, project_id)

        wb = load_workbook(BytesIO(result), data_only=False)
        pivot = wb["P&L Pivot"]

        # Header присутствует
        assert pivot.cell(row=1, column=1).value == PNL_PIVOT_HEADERS[0]
        assert pivot.max_column >= len(PNL_PIVOT_HEADERS)

        # 43 data rows + 1 header = 44 строки минимум
        assert pivot.max_row >= 44

        # Первая data row: SKU-поля не пустые
        assert pivot.cell(row=2, column=1).value == "Gorji"   # brand
        assert pivot.cell(row=2, column=2).value == "Calc test SKU"  # name
        # Период: первый период = M1
        assert pivot.cell(row=2, column=10).value == "M1"     # Период
        assert pivot.cell(row=2, column=11).value == "monthly"  # Тип периода
        assert pivot.cell(row=2, column=12).value == 1         # Год


# ============================================================
# Endpoint tests
# ============================================================


class TestExportXlsxEndpoint:
    async def test_returns_xlsx_with_correct_mime(
        self,
        auth_client: AsyncClient,
        db_session: AsyncSession,
    ):
        """GET /api/projects/{id}/export/xlsx → 200 + правильный MIME."""
        project_id, _, _, _ = await _seed_minimal_project(db_session)
        await db_session.commit()

        resp = await auth_client.get(
            f"/api/projects/{project_id}/export/xlsx"
        )
        assert resp.status_code == 200, resp.text
        assert resp.headers["content-type"].startswith(XLSX_MIME)

    async def test_response_is_valid_xlsx(
        self,
        auth_client: AsyncClient,
        db_session: AsyncSession,
    ):
        """Body open-able через openpyxl, 4 листа (Вводные/PnL/KPI/P&L Pivot)."""
        project_id, _, _, _ = await _seed_minimal_project(db_session)
        await db_session.commit()

        resp = await auth_client.get(
            f"/api/projects/{project_id}/export/xlsx"
        )
        assert resp.status_code == 200

        wb = load_workbook(BytesIO(resp.content), data_only=False)
        assert len(wb.sheetnames) == 4  # C #15: added P&L Pivot sheet
        assert "Вводные" in wb.sheetnames

    async def test_filename_in_content_disposition(
        self,
        auth_client: AsyncClient,
        db_session: AsyncSession,
    ):
        """Content-Disposition содержит filename с project_id."""
        project_id, _, _, _ = await _seed_minimal_project(db_session)
        await db_session.commit()

        resp = await auth_client.get(
            f"/api/projects/{project_id}/export/xlsx"
        )
        cd = resp.headers.get("content-disposition", "")
        assert "attachment" in cd
        assert f"project_{project_id}" in cd
        assert ".xlsx" in cd

    async def test_404_for_unknown_project(
        self, auth_client: AsyncClient
    ):
        resp = await auth_client.get("/api/projects/999999/export/xlsx")
        assert resp.status_code == 404

    async def test_unauthorized(self, client: AsyncClient):
        resp = await client.get("/api/projects/1/export/xlsx")
        assert resp.status_code == 401

    async def test_cyrillic_project_name_does_not_break_header(
        self,
        auth_client: AsyncClient,
        db_session: AsyncSession,
    ):
        """Regression: кириллическое имя проекта не ломает Content-Disposition.

        Был баг UnicodeEncodeError 'latin-1' — Python isalnum() возвращает
        True для кириллицы, она попадала в HTTP header и рушила экспорт.
        Исправлено через RFC 5987 filename*=UTF-8 + ASCII fallback.
        """
        from app.models import Project

        project_id, _, _, _ = await _seed_minimal_project(db_session)
        project = await db_session.get(Project, project_id)
        assert project is not None
        project.name = "Проект на русском"
        await db_session.flush()
        await db_session.commit()

        resp = await auth_client.get(
            f"/api/projects/{project_id}/export/xlsx"
        )
        assert resp.status_code == 200, resp.text
        cd = resp.headers.get("content-disposition", "")
        assert "filename=" in cd
        assert "filename*=UTF-8''" in cd
        assert "%D0" in cd
